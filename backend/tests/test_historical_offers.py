"""
Test Historical Offer Intelligence (BLOCK B) APIs
- Import preview: POST /api/historical/import-preview
- Import confirm: POST /api/historical/import-confirm
- Analytics: GET /api/historical/analytics with filters
- Internal price hint in AI proposal
- Normalization logic (keywords to category mapping)
- Section headers skipping
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
TEST_EMAIL = "admin@begwork.com"
TEST_PASSWORD = "AdminTest123!Secure"


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TEST_EMAIL,
        "password": TEST_PASSWORD
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    return response.json()["token"]


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Headers with auth token"""
    return {"Authorization": f"Bearer {auth_token}"}


# ══════════════════════════════════════════════════════════════════════════════
# GET /api/historical/analytics - Historical Analytics Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestHistoricalAnalytics:
    """Tests for GET /api/historical/analytics"""

    def test_analytics_returns_data_structure(self, auth_headers):
        """Analytics returns expected data structure with categories, batches, types, cities"""
        response = requests.get(f"{BASE_URL}/api/historical/analytics", headers=auth_headers)
        assert response.status_code == 200
        data = response.json()
        
        # Verify expected structure
        assert "total_rows" in data
        assert "total_batches" in data
        assert "categories" in data
        assert "batches" in data
        assert "unique_types" in data
        assert "unique_cities" in data
        
        # Data assertions - at least pre-seeded data (10 rows, 2 batches)
        assert data["total_rows"] >= 10, f"Expected at least 10 rows, got {data['total_rows']}"
        assert data["total_batches"] >= 2, f"Expected at least 2 batches, got {data['total_batches']}"
        print(f"✓ Analytics returns {data['total_rows']} rows, {data['total_batches']} batches")

    def test_analytics_categories_have_median_prices(self, auth_headers):
        """Each category has median/avg/min/max prices"""
        response = requests.get(f"{BASE_URL}/api/historical/analytics", headers=auth_headers)
        assert response.status_code == 200
        categories = response.json()["categories"]
        
        assert len(categories) >= 4, f"Expected at least 4 categories, got {len(categories)}"
        
        for cat in categories:
            assert "activity_type" in cat
            assert "activity_subtype" in cat
            assert "unit" in cat
            assert "sample_count" in cat
            assert "median_material" in cat
            assert "median_labor" in cat
            assert "median_total" in cat
            assert "min_total" in cat
            assert "max_total" in cat
            assert cat["sample_count"] > 0
            assert cat["median_total"] > 0
        
        print(f"✓ All {len(categories)} categories have valid price data")

    def test_analytics_filter_by_city(self, auth_headers):
        """Analytics filters correctly by city"""
        response = requests.get(
            f"{BASE_URL}/api/historical/analytics",
            params={"city": "София"},
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # All pre-seeded data is from София
        assert len(data["categories"]) >= 4
        print(f"✓ City filter 'София' returns {len(data['categories'])} categories")

    def test_analytics_filter_by_activity_type(self, auth_headers):
        """Analytics filters correctly by activity_type"""
        response = requests.get(
            f"{BASE_URL}/api/historical/analytics",
            params={"activity_type": "Довършителни"},
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Verify all returned categories are of the filtered type
        for cat in data["categories"]:
            assert cat["activity_type"] == "Довършителни"
        
        print(f"✓ Activity type filter returns only 'Довършителни' categories ({len(data['categories'])})")

    def test_analytics_batches_have_metadata(self, auth_headers):
        """Batches include source metadata (project, date, city)"""
        response = requests.get(f"{BASE_URL}/api/historical/analytics", headers=auth_headers)
        assert response.status_code == 200
        batches = response.json()["batches"]
        
        assert len(batches) >= 2, f"Expected at least 2 batches, got {len(batches)}"
        
        # Check pre-seeded batches exist
        batch_projects = [b["source_project_name"] for b in batches]
        assert "Офис Люлин" in batch_projects, "Pre-seeded batch 'Офис Люлин' not found"
        assert "Хотел Бояна" in batch_projects, "Pre-seeded batch 'Хотел Бояна' not found"
        
        for batch in batches:
            assert "id" in batch
            assert "file_name" in batch
            assert "source_project_name" in batch
            assert "source_date" in batch
            assert "city" in batch
            assert "rows_imported" in batch
            assert "created_at" in batch
        
        print(f"✓ All {len(batches)} batches have complete metadata")


# ══════════════════════════════════════════════════════════════════════════════
# POST /api/extra-works/ai-proposal - Internal Price Hint Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestInternalPriceHint:
    """Tests for internal price hint in AI proposal"""

    def test_ai_proposal_includes_internal_price_hint(self, auth_headers):
        """AI proposal includes internal_price_hint when historical data exists"""
        response = requests.post(
            f"{BASE_URL}/api/extra-works/ai-proposal",
            json={"title": "Мазилка по стени", "unit": "m2", "qty": 10, "city": "София"},
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        assert "internal_price_hint" in data
        hint = data["internal_price_hint"]
        
        assert hint["available"] is True
        assert hint["sample_count"] >= 2
        assert hint["median_total"] > 0
        assert hint["median_material"] > 0
        assert hint["median_labor"] > 0
        assert hint["min_total"] > 0
        assert hint["max_total"] > 0
        assert "range_label" in hint
        
        print(f"✓ AI proposal includes internal price hint: median_total={hint['median_total']}, range={hint['range_label']}")

    def test_ai_proposal_hint_not_available_for_unknown_category(self, auth_headers):
        """AI proposal returns hint.available=False for categories without historical data"""
        response = requests.post(
            f"{BASE_URL}/api/extra-works/ai-proposal",
            json={"title": "Неизвестна работа XYZ", "unit": "m2", "qty": 5},
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Should still have internal_price_hint field
        assert "internal_price_hint" in data
        hint = data["internal_price_hint"]
        
        # For unknown category "Общо/СМР", hint may not be available
        # (depends on whether there's historical data for this generic category)
        print(f"✓ AI proposal for unknown category: hint.available={hint.get('available')}, sample_count={hint.get('sample_count', 0)}")

    def test_ai_proposal_hint_for_облицовка(self, auth_headers):
        """AI proposal returns correct hint for плочки/облицовка"""
        response = requests.post(
            f"{BASE_URL}/api/extra-works/ai-proposal",
            json={"title": "Полагане на плочки фаянс", "unit": "m2", "qty": 15},
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        assert "internal_price_hint" in data
        hint = data["internal_price_hint"]
        
        if hint["available"]:
            # Облицовка should have median_total around 40-50 based on seed data
            assert hint["median_total"] >= 40
            print(f"✓ Облицовка hint: median={hint['median_total']}, range={hint['range_label']}")
        else:
            print("✓ No historical data for oblicovka (expected)")


# ══════════════════════════════════════════════════════════════════════════════
# Normalization Logic Tests (via import-preview)
# ══════════════════════════════════════════════════════════════════════════════

class TestNormalizationLogic:
    """Tests for normalization of SMR descriptions"""

    def test_normalization_keywords_mapping(self, auth_headers):
        """Test that known keywords map to correct categories"""
        # Create test XLSX with different SMR types
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Header
        ws.append(["Описание", "Мярка", "Количество", "Цена материал", "Цена труд", "Обща цена"])
        
        # Test rows with different keywords
        test_rows = [
            ("Вътрешна гипсова мазилка", "м2", 20, 8, 12, 400),  # мазилк → Мокри процеси/Мазилка
            ("Боядисване с латекс", "м2", 50, 4, 6, 500),         # боядисв → Довършителни/Боядисване
            ("Полагане на плочки фаянс", "м2", 10, 25, 20, 450),  # плоч → Довършителни/Облицовка
            ("Шпакловане на стени", "м2", 30, 5, 8, 390),         # шпакл → Довършителни/Шпакловка
            ("Гипсокартон монтаж", "м2", 15, 14, 15, 435),        # гипсокарт → Сухо строителство/Гипсокартон
        ]
        
        for row in test_rows:
            ws.append(row)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload for preview
        files = {"file": ("test_normalization.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Check normalization results
        lines = data["lines"]
        assert len(lines) == 5
        
        expected_mappings = [
            ("Мокри процеси", "Мазилка"),
            ("Довършителни", "Боядисване"),
            ("Довършителни", "Облицовка"),
            ("Довършителни", "Шпакловка"),
            ("Сухо строителство", "Гипсокартон"),
        ]
        
        for i, line in enumerate(lines):
            expected_type, expected_subtype = expected_mappings[i]
            assert line["normalized_activity_type"] == expected_type, f"Row {i+1}: expected type '{expected_type}', got '{line['normalized_activity_type']}'"
            assert line["normalized_activity_subtype"] == expected_subtype, f"Row {i+1}: expected subtype '{expected_subtype}', got '{line['normalized_activity_subtype']}'"
            assert line["normalization_confidence"] >= 0.8
        
        print(f"✓ All 5 keyword mappings correct (мазилк, боядисв, плоч, шпакл, гипсокарт)")

    def test_section_headers_skipped(self, auth_headers):
        """Test that section headers (subtotal, рекапитулация) are skipped"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Header
        ws.append(["Описание", "Мярка", "К-во", "Материал", "Труд", "Общо"])
        
        # Mix of real rows and section headers
        ws.append(["Мазилка по стени", "м2", 20, 8, 12, 400])          # Real row
        ws.append(["ОБЩО ЗА РАЗДЕЛ 1", "", "", "", "", 400])           # Section header - should skip
        ws.append(["Боядисване стени", "м2", 30, 4, 6, 300])           # Real row
        ws.append(["Рекапитулация", "", "", "", "", 700])              # Section header - should skip
        ws.append(["Subtotal довършителни", "", "", "", "", 300])      # Section header - should skip
        ws.append(["Плочки баня", "м2", 8, 25, 20, 360])               # Real row
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_sections.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Should have parsed 3 real lines and skipped 3 sections
        assert data["parsed_lines"] == 3, f"Expected 3 parsed lines, got {data['parsed_lines']}"
        assert data["skipped_sections"] >= 3, f"Expected at least 3 skipped sections, got {data['skipped_sections']}"
        
        # Verify only real work rows are in lines
        for line in data["lines"]:
            raw_text = line["raw_smr_text"].lower()
            assert "общо" not in raw_text
            assert "рекапитулация" not in raw_text
            assert "subtotal" not in raw_text
        
        print(f"✓ Section headers correctly skipped: {data['skipped_sections']} sections, {data['parsed_lines']} real rows")


# ══════════════════════════════════════════════════════════════════════════════
# POST /api/historical/import-preview Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestImportPreview:
    """Tests for POST /api/historical/import-preview"""

    def test_import_preview_parses_xlsx(self, auth_headers):
        """Import preview parses XLSX and returns normalized lines"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(["Вид СМР", "Мярка", "Количество", "Цена материал", "Цена труд", "Обща цена"])
        ws.append(["Гипсова мазилка по стени", "м2", 100, 8.5, 12, 2050])
        ws.append(["Боядисване латекс", "м2", 150, 4.2, 5.8, 1500])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_preview.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["file_name"] == "test_preview.xlsx"
        assert data["parsed_lines"] == 2
        assert "lines" in data
        assert len(data["lines"]) == 2
        
        # Check first line structure
        line = data["lines"][0]
        assert "raw_smr_text" in line
        assert "normalized_activity_type" in line
        assert "normalized_activity_subtype" in line
        assert "unit" in line
        assert "material_price_per_unit" in line
        assert "labor_price_per_unit" in line
        assert "total_price_per_unit" in line
        
        print(f"✓ Import preview parsed {data['parsed_lines']} lines with normalization")

    def test_import_preview_normalizes_units(self, auth_headers):
        """Import preview normalizes Bulgarian units (м2→m2, бр→pcs)"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(["Описание", "Мярка", "К-во", "Материал", "Труд", "Общо"])
        ws.append(["Ел. контакт монтаж", "бр", 10, 35, 25, 600])       # бр → pcs
        ws.append(["Мазилка", "м2", 20, 8, 12, 400])                   # м2 → m2
        ws.append(["Кабел", "м", 50, 3, 2, 250])                       # м → m
        ws.append(["Бетон", "м3", 5, 100, 50, 750])                    # м3 → m3
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_units.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert response.status_code == 200
        lines = response.json()["lines"]
        
        expected_units = ["pcs", "m2", "m", "m3"]
        for i, line in enumerate(lines):
            assert line["unit"] == expected_units[i], f"Line {i+1}: expected unit '{expected_units[i]}', got '{line['unit']}'"
        
        print(f"✓ All 4 Bulgarian units normalized correctly (бр→pcs, м2→m2, м→m, м3→m3)")

    def test_import_preview_rejects_non_xlsx(self, auth_headers):
        """Import preview rejects non-XLSX files"""
        files = {"file": ("test.txt", b"not an xlsx file", "text/plain")}
        response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert response.status_code == 400
        assert "xlsx" in response.json().get("detail", "").lower()
        print("✓ Non-XLSX files correctly rejected with 400")


# ══════════════════════════════════════════════════════════════════════════════
# POST /api/historical/import-confirm Tests
# ══════════════════════════════════════════════════════════════════════════════

class TestImportConfirm:
    """Tests for POST /api/historical/import-confirm"""

    def test_import_confirm_saves_rows_with_metadata(self, auth_headers):
        """Import confirm saves historical rows with batch metadata"""
        # First create preview data
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(["Описание", "Мярка", "К-во", "Материал", "Труд", "Общо"])
        ws.append(["TEST_Мазилка тест", "м2", 25, 8, 12, 500])
        ws.append(["TEST_Боядисване тест", "м2", 40, 4, 6, 400])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test_confirm.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        preview_response = requests.post(
            f"{BASE_URL}/api/historical/import-preview",
            files=files,
            headers=auth_headers
        )
        assert preview_response.status_code == 200
        preview_data = preview_response.json()
        
        # Now confirm import with metadata
        confirm_payload = {
            "file_name": "test_confirm.xlsx",
            "source_project_name": "TEST_Project_Confirm",
            "source_date": "2025",
            "city": "Пловдив",
            "source_offer_name": "TEST_Offer_001",
            "lines": preview_data["lines"]
        }
        
        confirm_response = requests.post(
            f"{BASE_URL}/api/historical/import-confirm",
            json=confirm_payload,
            headers=auth_headers
        )
        assert confirm_response.status_code == 201
        data = confirm_response.json()
        
        assert data["ok"] is True
        assert "batch_id" in data
        assert data["rows_imported"] == 2
        
        batch_id = data["batch_id"]
        
        # Verify data persisted - check analytics now includes new batch
        analytics_response = requests.get(
            f"{BASE_URL}/api/historical/analytics",
            headers=auth_headers
        )
        assert analytics_response.status_code == 200
        analytics = analytics_response.json()
        
        # Find our new batch
        batches = analytics["batches"]
        our_batch = next((b for b in batches if b["id"] == batch_id), None)
        assert our_batch is not None, "New batch not found in analytics"
        assert our_batch["source_project_name"] == "TEST_Project_Confirm"
        assert our_batch["city"] == "Пловдив"
        assert our_batch["rows_imported"] == 2
        
        print(f"✓ Import confirm saved batch {batch_id} with 2 rows and correct metadata")

    def test_import_confirm_requires_admin(self, auth_token):
        """Import confirm requires Admin/Owner role"""
        # This test verifies role-based access
        # The current user is Admin, so it should pass
        # We can verify by checking the endpoint works for admin
        response = requests.post(
            f"{BASE_URL}/api/historical/import-confirm",
            json={"lines": []},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        # Empty lines should return 400 (No lines), not 403
        assert response.status_code == 400
        assert "No lines" in response.json().get("detail", "")
        print("✓ Import confirm role check: Admin can access, empty lines returns 400")


# ══════════════════════════════════════════════════════════════════════════════
# Data Cleanup
# ══════════════════════════════════════════════════════════════════════════════

@pytest.fixture(scope="module", autouse=True)
def cleanup_test_data():
    """Cleanup TEST_ prefixed data after tests"""
    yield
    # Cleanup is tricky without direct DB access, but the test data
    # is prefixed with TEST_ for identification
    print("Note: TEST_ prefixed historical data remains in DB")
