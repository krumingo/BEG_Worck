"""
Tests for Excel Import V2.
Run: pytest backend/tests/test_excel_import_v2.py -v
"""
import pytest
import httpx
import io
from pathlib import Path
from openpyxl import Workbook

BASE_URL = ""
fe_env = Path(__file__).parent.parent.parent / "frontend" / ".env"
if fe_env.exists():
    for line in fe_env.read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
            break

API = f"{BASE_URL}/api"
EMAIL = "admin@begwork.com"
PASSWORD = "AdminTest123!Secure"


@pytest.fixture(scope="module")
def auth_headers():
    r = httpx.post(f"{API}/auth/login", json={"email": EMAIL, "password": PASSWORD})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


@pytest.fixture(scope="module")
def project_id(auth_headers):
    r = httpx.get(f"{API}/projects", headers=auth_headers)
    projects = r.json() if isinstance(r.json(), list) else r.json()
    return projects[0]["id"]


def make_xlsx(rows, sheet_name="КСС"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Test 1: Detect structure ───────────────────────────────────────

def test_detect_structure(auth_headers):
    xlsx = make_xlsx([
        ["№", "Описание на дейност", "Мерна ед.", "Количество", "Цена материал", "Цена труд", "Общо"],
        [1, "Шпакловка", "м2", 50, 3, 8, 550],
        [2, "Боядисване", "м2", 50, 2, 6, 400],
    ])
    r = httpx.post(f"{API}/excel-import/preview",
        files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"import_type": "kss"},
        headers=auth_headers,
    )
    assert r.status_code == 200
    d = r.json()
    assert "detected_columns" in d
    assert "smr_type" in d["detected_columns"]
    assert d["confidence"] > 0
    assert len(d["preview_rows"]) == 2


# ── Test 2: Multi-sheet ────────────────────────────────────────────

def test_multi_sheet(auth_headers):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Лист1"
    ws1.append(["Текст"])
    ws2 = wb.create_sheet("КСС")
    ws2.append(["№", "Описание", "Ед.", "Кол-во"])
    ws2.append([1, "Мазилка", "м2", 30])
    buf = io.BytesIO()
    wb.save(buf)

    r = httpx.post(f"{API}/excel-import/preview",
        files={"file": ("multi.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"import_type": "kss", "sheet_name": "КСС"},
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert "КСС" in r.json()["sheet_names"]
    assert r.json()["selected_sheet"] == "КСС"


# ── Test 3: Preview with warnings ──────────────────────────────────

def test_preview_warnings(auth_headers):
    xlsx = make_xlsx([
        ["Номер", "Нещо", "Стойност"],
        [1, "Ред1", 100],
    ])
    r = httpx.post(f"{API}/excel-import/preview",
        files={"file": ("warn.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={},
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert len(r.json()["warnings"]) >= 1


# ── Test 4: Save template ──────────────────────────────────────────

def test_save_template(auth_headers):
    r = httpx.post(f"{API}/excel-import/templates", json={
        "name": "Стандартен КСС",
        "import_type": "kss",
        "column_mapping": {"smr_type": "B", "unit": "C", "qty": "D", "total_price": "G"},
    }, headers=auth_headers)
    assert r.status_code == 201
    assert r.json()["name"] == "Стандартен КСС"


# ── Test 5: Apply template preview ─────────────────────────────────

def test_apply_template(auth_headers):
    # Get template
    templates = httpx.get(f"{API}/excel-import/templates", headers=auth_headers).json()["items"]
    assert len(templates) >= 1
    tid = templates[0]["id"]

    xlsx = make_xlsx([
        ["№", "Описание", "Ед.", "Кол-во", "Мат.", "Труд", "Общо"],
        [1, "Плочки", "м2", 20, 15, 22, 740],
    ])
    r = httpx.post(f"{API}/excel-import/preview-with-template",
        files={"file": ("tpl.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"template_id": tid},
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert "template" in r.json()
    assert r.json()["total"] >= 1


# ── Test 6: Update template ────────────────────────────────────────

def test_update_template(auth_headers):
    templates = httpx.get(f"{API}/excel-import/templates", headers=auth_headers).json()["items"]
    tid = templates[0]["id"]
    r = httpx.put(f"{API}/excel-import/templates/{tid}", json={"notes": "Updated"}, headers=auth_headers)
    assert r.status_code == 200
    assert r.json()["notes"] == "Updated"


# ── Test 7: Delete template ────────────────────────────────────────

def test_delete_template(auth_headers):
    t = httpx.post(f"{API}/excel-import/templates", json={
        "name": "Delete Me", "import_type": "kss", "column_mapping": {"smr_type": "A"},
    }, headers=auth_headers).json()
    r = httpx.delete(f"{API}/excel-import/templates/{t['id']}", headers=auth_headers)
    assert r.status_code == 200


# ── Test 8: Commit bridge ──────────────────────────────────────────

def test_commit(auth_headers, project_id):
    xlsx = make_xlsx([
        ["№", "Описание на дейност", "Мерна ед.", "Количество", "Материал", "Труд"],
        [1, "Грундиране", "м2", 40, 1.5, 3],
        [2, "Латекс", "м2", 40, 2, 5],
    ])
    r = httpx.post(f"{API}/excel-import/commit",
        files={"file": ("commit.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"project_id": project_id, "import_type": "kss", "name": "V2 Commit Test"},
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert r.json()["lines_imported"] == 2
    assert r.json()["analysis_id"] is not None
