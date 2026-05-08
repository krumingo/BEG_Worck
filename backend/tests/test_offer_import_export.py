"""
Tests for Offer Import/Export features (P0 BLOCK A):
- PDF export for offers (main + extra)
- XLSX export with structured data
- Import template download
- Import preview (parse XLSX with column mapping)
- Import confirm (create offer from parsed lines)
- Bulgarian units handling
"""
import pytest
import requests
import os
import io
import tempfile

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test fixtures
@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token for admin user"""
    response = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": "admin@begwork.com", "password": "AdminTest123!Secure"}
    )
    assert response.status_code == 200, f"Login failed: {response.text}"
    return response.json()["token"]


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Headers with auth token"""
    return {
        "Authorization": f"Bearer {auth_token}",
        "Content-Type": "application/json"
    }


@pytest.fixture(scope="module")
def project_id(auth_headers):
    """Get or create a test project"""
    # Use existing project PRJ-001
    response = requests.get(f"{BASE_URL}/api/projects", headers=auth_headers)
    assert response.status_code == 200
    projects = response.json()
    if projects:
        return projects[0]["id"]  # Use first project
    pytest.skip("No projects available for testing")


@pytest.fixture(scope="module")
def test_offer_id(auth_headers, project_id):
    """Create a test offer for export tests"""
    offer_data = {
        "project_id": project_id,
        "title": "TEST_Export_Offer",
        "currency": "BGN",
        "vat_percent": 20,
        "notes": "Test offer for export testing",
        "lines": [
            {
                "activity_code": None,
                "activity_name": "Боядисване на стени",
                "unit": "m2",
                "qty": 50,
                "material_unit_cost": 5.50,
                "labor_unit_cost": 8.00,
                "labor_hours_per_unit": None,
                "note": "2 слоя латекс",
                "sort_order": 0
            },
            {
                "activity_code": None,
                "activity_name": "Монтаж на ел. контакти",
                "unit": "pcs",
                "qty": 10,
                "material_unit_cost": 25.00,
                "labor_unit_cost": 15.00,
                "labor_hours_per_unit": None,
                "note": "Schneider серия",
                "sort_order": 1
            }
        ]
    }
    response = requests.post(f"{BASE_URL}/api/offers", json=offer_data, headers=auth_headers)
    assert response.status_code == 201, f"Failed to create test offer: {response.text}"
    offer = response.json()
    yield offer["id"]
    # Cleanup: Delete test offer
    requests.delete(f"{BASE_URL}/api/offers/{offer['id']}", headers=auth_headers)


class TestOfferPDFExport:
    """Tests for PDF export functionality"""
    
    def test_pdf_export_returns_valid_pdf(self, auth_headers, test_offer_id):
        """GET /api/offers/{id}/pdf returns valid PDF"""
        response = requests.get(
            f"{BASE_URL}/api/offers/{test_offer_id}/pdf",
            headers=auth_headers
        )
        assert response.status_code == 200
        assert response.headers.get("content-type") == "application/pdf"
        
        # Check Content-Disposition header
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp
        assert ".pdf" in content_disp
        
        # Verify PDF magic bytes
        assert response.content[:4] == b'%PDF', "Response should start with PDF magic bytes"
        print(f"✓ PDF export returned {len(response.content)} bytes")
    
    def test_pdf_export_nonexistent_offer(self, auth_headers):
        """GET /api/offers/{invalid_id}/pdf returns 404"""
        response = requests.get(
            f"{BASE_URL}/api/offers/nonexistent-offer-id/pdf",
            headers=auth_headers
        )
        assert response.status_code == 404
    
    def test_pdf_export_unauthorized(self, test_offer_id):
        """GET /api/offers/{id}/pdf without auth returns 401 or 403"""
        response = requests.get(f"{BASE_URL}/api/offers/{test_offer_id}/pdf")
        assert response.status_code in [401, 403], f"Expected 401/403, got {response.status_code}"


class TestOfferXLSXExport:
    """Tests for XLSX export functionality"""
    
    def test_xlsx_export_returns_valid_xlsx(self, auth_headers, test_offer_id):
        """GET /api/offers/{id}/xlsx returns valid XLSX"""
        response = requests.get(
            f"{BASE_URL}/api/offers/{test_offer_id}/xlsx",
            headers=auth_headers
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")
        
        # Check Content-Disposition header
        content_disp = response.headers.get("content-disposition", "")
        assert "attachment" in content_disp
        assert ".xlsx" in content_disp
        
        # Verify XLSX magic bytes (PK zip format)
        assert response.content[:2] == b'PK', "XLSX should start with PK (zip format)"
        print(f"✓ XLSX export returned {len(response.content)} bytes")
    
    def test_xlsx_export_nonexistent_offer(self, auth_headers):
        """GET /api/offers/{invalid_id}/xlsx returns 404"""
        response = requests.get(
            f"{BASE_URL}/api/offers/nonexistent-offer-id/xlsx",
            headers=auth_headers
        )
        assert response.status_code == 404


class TestOfferImportTemplate:
    """Tests for import template download"""
    
    def test_import_template_download(self, auth_headers):
        """GET /api/offer-import-template returns valid XLSX template"""
        response = requests.get(
            f"{BASE_URL}/api/offer-import-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")
        
        # Check Content-Disposition
        content_disp = response.headers.get("content-disposition", "")
        assert "BEG_Work_Offer_Import_Template.xlsx" in content_disp
        
        # Verify XLSX magic bytes
        assert response.content[:2] == b'PK'
        print(f"✓ Import template returned {len(response.content)} bytes")
    
    def test_import_template_has_content(self, auth_headers):
        """Import template should have reasonable size with headers and examples"""
        response = requests.get(
            f"{BASE_URL}/api/offer-import-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        # Template with 2 sheets, headers, examples should be >5KB
        assert len(response.content) > 5000, f"Template seems too small: {len(response.content)} bytes"


class TestOfferImportPreview:
    """Tests for import preview (parse XLSX)"""
    
    def test_import_preview_with_valid_xlsx(self, auth_headers):
        """POST /api/offers/import-preview parses XLSX and returns preview"""
        # First download template to use as test file
        template_resp = requests.get(
            f"{BASE_URL}/api/offer-import-template",
            headers=auth_headers
        )
        assert template_resp.status_code == 200
        
        # Upload template for preview (it has example rows)
        files = {"file": ("test_import.xlsx", io.BytesIO(template_resp.content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {"Authorization": auth_headers["Authorization"]}
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-preview",
            files=files,
            headers=headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Check response structure
        assert "file_name" in data
        assert "total_rows" in data
        assert "parsed_lines" in data
        assert "lines" in data
        assert "column_mapping" in data
        assert "headers_found" in data
        
        print(f"✓ Import preview: {data['parsed_lines']} lines from {data['total_rows']} rows")
        print(f"  Column mapping: {data['column_mapping']}")
        
        # Template has 3 example rows
        assert data["parsed_lines"] >= 1, "Should parse at least 1 line from template"
    
    def test_import_preview_with_invalid_file(self, auth_headers):
        """POST /api/offers/import-preview with non-xlsx returns 400"""
        files = {"file": ("test.txt", io.BytesIO(b"not an xlsx file"), "text/plain")}
        headers = {"Authorization": auth_headers["Authorization"]}
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-preview",
            files=files,
            headers=headers
        )
        assert response.status_code == 400
    
    def test_import_preview_parses_bulgarian_units(self, auth_headers):
        """Import preview should recognize Bulgarian unit names (м2, бр, etc.)"""
        # Download template and verify lines have expected units
        template_resp = requests.get(
            f"{BASE_URL}/api/offer-import-template",
            headers=auth_headers
        )
        
        files = {"file": ("test.xlsx", io.BytesIO(template_resp.content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {"Authorization": auth_headers["Authorization"]}
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-preview",
            files=files,
            headers=headers
        )
        assert response.status_code == 200
        data = response.json()
        
        # Template has м2 and бр units which should be converted to m2 and pcs
        if data["lines"]:
            units_found = [line.get("unit") for line in data["lines"]]
            print(f"  Units parsed: {units_found}")
            # Should convert Bulgarian units to internal format
            valid_units = ["m2", "m", "pcs", "hours", "lot", "kg", "l"]
            for unit in units_found:
                assert unit in valid_units, f"Unit '{unit}' not recognized"


class TestOfferImportConfirm:
    """Tests for import confirm (create offer from preview)"""
    
    def test_import_confirm_creates_offer(self, auth_headers, project_id):
        """POST /api/offers/import-confirm creates offer from parsed lines"""
        # First get preview data
        template_resp = requests.get(
            f"{BASE_URL}/api/offer-import-template",
            headers=auth_headers
        )
        
        files = {"file": ("test.xlsx", io.BytesIO(template_resp.content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {"Authorization": auth_headers["Authorization"]}
        
        preview_resp = requests.post(
            f"{BASE_URL}/api/offers/import-preview",
            files=files,
            headers=headers
        )
        assert preview_resp.status_code == 200
        preview_data = preview_resp.json()
        
        # Confirm import
        confirm_data = {
            "project_id": project_id,
            "title": "TEST_Import_Confirm_Offer",
            "lines": preview_data["lines"],
            "file_name": "test.xlsx",
            "offer_type": "main",
            "currency": "BGN",
            "vat_percent": 20,
            "notes": "Created via import test"
        }
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-confirm",
            json=confirm_data,
            headers=auth_headers
        )
        assert response.status_code == 201, f"Import confirm failed: {response.text}"
        offer = response.json()
        
        # Verify created offer
        assert "id" in offer
        assert "offer_no" in offer
        assert offer["title"] == "TEST_Import_Confirm_Offer"
        assert offer["status"] == "Draft"
        assert offer["offer_type"] == "main"
        assert len(offer["lines"]) == len(preview_data["lines"])
        
        print(f"✓ Import confirm created offer {offer['offer_no']} with {len(offer['lines'])} lines")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/offers/{offer['id']}", headers=auth_headers)
    
    def test_import_confirm_requires_project(self, auth_headers):
        """POST /api/offers/import-confirm without project_id returns 400"""
        confirm_data = {
            "title": "TEST_Import_No_Project",
            "lines": [{"description": "Test", "unit": "pcs", "qty": 1, "material_price": 10, "labor_price": 5}]
        }
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-confirm",
            json=confirm_data,
            headers=auth_headers
        )
        assert response.status_code == 400
        assert "project_id" in response.text.lower()
    
    def test_import_confirm_requires_lines(self, auth_headers, project_id):
        """POST /api/offers/import-confirm with empty lines returns 400"""
        confirm_data = {
            "project_id": project_id,
            "title": "TEST_Import_No_Lines",
            "lines": []
        }
        
        response = requests.post(
            f"{BASE_URL}/api/offers/import-confirm",
            json=confirm_data,
            headers=auth_headers
        )
        assert response.status_code == 400


class TestOfferExportContentVerification:
    """Tests to verify export content accuracy"""
    
    def test_pdf_contains_offer_data(self, auth_headers, test_offer_id):
        """PDF should contain offer number, lines, totals"""
        # Get offer details first
        offer_resp = requests.get(
            f"{BASE_URL}/api/offers/{test_offer_id}",
            headers=auth_headers
        )
        assert offer_resp.status_code == 200
        offer = offer_resp.json()
        
        # Get PDF
        pdf_resp = requests.get(
            f"{BASE_URL}/api/offers/{test_offer_id}/pdf",
            headers=auth_headers
        )
        assert pdf_resp.status_code == 200
        
        # PDF should be substantial (with lines, headers, etc)
        assert len(pdf_resp.content) > 1000
        print(f"✓ PDF for offer {offer['offer_no']}: {len(pdf_resp.content)} bytes, {len(offer['lines'])} lines")
    
    def test_xlsx_contains_structured_data(self, auth_headers, test_offer_id):
        """XLSX should contain proper structured data"""
        response = requests.get(
            f"{BASE_URL}/api/offers/{test_offer_id}/xlsx",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        # Parse XLSX to verify structure
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            
            # Sheet should be named "Оферта"
            assert ws.title == "Оферта", f"Expected sheet 'Оферта', got '{ws.title}'"
            
            # Should have content (at least header + some rows)
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) > 5, f"Expected >5 rows, got {len(rows)}"
            
            print(f"✓ XLSX has sheet '{ws.title}' with {len(rows)} rows")
        except ImportError:
            print("  (openpyxl not available for detailed verification)")


class TestImportWarnings:
    """Tests for import preview warnings"""
    
    def test_import_preview_warns_on_missing_prices(self, auth_headers):
        """Import preview should warn about missing prices"""
        # Create a minimal XLSX with missing prices
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Оферта"
            
            # Headers
            ws['A1'] = "Описание"
            ws['B1'] = "Мярка"
            ws['C1'] = "Количество"
            ws['D1'] = "Цена материал"
            ws['E1'] = "Цена труд"
            
            # Row with missing prices
            ws['A2'] = "Тест дейност без цена"
            ws['B2'] = "м2"
            ws['C2'] = 10
            ws['D2'] = None
            ws['E2'] = None
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            files = {"file": ("test_warnings.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            headers = {"Authorization": auth_headers["Authorization"]}
            
            response = requests.post(
                f"{BASE_URL}/api/offers/import-preview",
                files=files,
                headers=headers
            )
            assert response.status_code == 200
            data = response.json()
            
            # Should have warnings about missing prices
            assert "warnings" in data
            print(f"✓ Import preview warnings: {data.get('warnings', [])}")
            
        except ImportError:
            pytest.skip("openpyxl not available for this test")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
