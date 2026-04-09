"""
Service - Excel Import/Export for KSS (Количествено-стойностна сметка).
Smart column detection + preview + multi-sheet support.
"""
import io
import uuid
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


def parse_float(val) -> float:
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


# ── Smart Column Detection ─────────────────────────────────────────

COLUMN_KEYWORDS = {
    "smr_type": ["описание", "дейност", "наименование", "смр", "работа", "description", "activity", "item", "вид"],
    "unit": ["мярка", "мерна", "ед.", "единица", "unit", "мерна ед"],
    "qty": ["количество", "кол.", "к-во", "qty", "quantity", "бр."],
    "material_price": ["материал", "мат. цена", "цена мат", "material", "мат./ед"],
    "labor_price": ["труд", "цена труд", "работна", "labor", "труд/ед"],
    "total_price": ["обща", "стойност", "общо", "total", "сума", "amount", "общо лв"],
    "number": ["№", "но.", "номер", "#", "num", "n."],
}


def detect_columns(ws, header_row=1) -> dict:
    """Detect column mapping from header keywords."""
    mapping = {}
    header_cells = []

    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        header_cells = row
        break

    for cell in header_cells:
        val = str(cell.value or "").lower().strip()
        if not val:
            continue
        col_letter = get_column_letter(cell.column)
        for field, keywords in COLUMN_KEYWORDS.items():
            if field in mapping:
                continue
            for kw in keywords:
                if kw in val:
                    mapping[field] = col_letter
                    break

    return mapping


def detect_header_row(ws) -> int:
    """Find the header row by looking for keyword matches."""
    for row_num in range(1, 8):
        matches = 0
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            for val in row:
                v = str(val or "").lower()
                for keywords in COLUMN_KEYWORDS.values():
                    if any(kw in v for kw in keywords):
                        matches += 1
                        break
        if matches >= 2:
            return row_num
    return 1


def col_letter_to_index(letter: str) -> int:
    """Convert column letter to 0-based index."""
    letter = letter.upper()
    result = 0
    for c in letter:
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1


def get_cell_by_mapping(row_values, mapping, field, fallback_index=None):
    """Get cell value using column mapping or fallback index."""
    if field in mapping:
        idx = col_letter_to_index(mapping[field])
        if idx < len(row_values):
            return row_values[idx]
    if fallback_index is not None and fallback_index < len(row_values):
        return row_values[fallback_index]
    return None


# ── Preview ────────────────────────────────────────────────────────

def preview_excel(file_bytes: bytes, sheet_name: str = None) -> dict:
    """Parse Excel and return preview without saving."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames

    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb.active

    header_row = detect_header_row(ws)
    mapping = detect_columns(ws, header_row)
    data_start = header_row + 1

    warnings = []
    if "smr_type" not in mapping:
        warnings.append("Колона 'описание/дейност' не е намерена — ще се ползва колона B")
    if "qty" not in mapping:
        warnings.append("Колона 'количество' не е намерена — ще бъде 1 по подразбиране")
    if "material_price" not in mapping and "total_price" not in mapping:
        warnings.append("Колона 'цена' не е намерена — цените ще бъдат празни")

    preview_rows = []
    total_rows = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue
        smr = get_cell_by_mapping(row, mapping, "smr_type", 1)
        if not smr:
            continue
        total_rows += 1
        if len(preview_rows) < 10:
            preview_rows.append({
                "smr_type": str(smr).strip(),
                "unit": str(get_cell_by_mapping(row, mapping, "unit", 2) or "m2").strip(),
                "qty": parse_float(get_cell_by_mapping(row, mapping, "qty", 3)),
                "material_price": parse_float(get_cell_by_mapping(row, mapping, "material_price", 4)),
                "labor_price": parse_float(get_cell_by_mapping(row, mapping, "labor_price", 5)),
                "total_price": parse_float(get_cell_by_mapping(row, mapping, "total_price", 6)),
            })

    return {
        "detected_columns": mapping,
        "header_row": header_row,
        "preview_rows": preview_rows,
        "total_rows": total_rows,
        "warnings": warnings,
        "sheet_names": sheet_names,
    }


# ── Import ─────────────────────────────────────────────────────────

async def import_kss_from_excel(
    file_bytes: bytes, org_id: str, project_id: str, created_by: str,
    project_name: str = "", column_mapping: dict = None, sheet_name: str = None
) -> dict:
    """Parse Excel file and return KSS lines with smart column detection."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    header_row = detect_header_row(ws)
    mapping = column_mapping or detect_columns(ws, header_row)
    data_start = header_row + 1

    lines = []
    warnings = []
    skipped = 0
    line_order = 0

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or all(v is None for v in row):
            skipped += 1
            continue

        smr_type = get_cell_by_mapping(row, mapping, "smr_type", 1)
        if not smr_type:
            skipped += 1
            continue
        smr_type = str(smr_type).strip()

        unit = str(get_cell_by_mapping(row, mapping, "unit", 2) or "m2").strip()
        qty = parse_float(get_cell_by_mapping(row, mapping, "qty", 3))
        mat_price = parse_float(get_cell_by_mapping(row, mapping, "material_price", 4))
        labor_price = parse_float(get_cell_by_mapping(row, mapping, "labor_price", 5))
        total_price = parse_float(get_cell_by_mapping(row, mapping, "total_price", 6))

        # Infer prices if only total is given
        if total_price > 0 and mat_price == 0 and labor_price == 0 and qty > 0:
            per_unit = total_price / qty
            mat_price = round(per_unit * 0.4, 2)
            labor_price = round(per_unit * 0.6, 2)
            warnings.append(f"Ред {line_order + 1} ({smr_type}): цените са разпределени автоматично (40/60)")

        if qty <= 0:
            qty = 1
            warnings.append(f"Ред {line_order + 1} ({smr_type}): количество зададено на 1")

        line = {
            "line_id": str(uuid.uuid4()),
            "smr_type": smr_type,
            "smr_subtype": "",
            "unit": unit,
            "qty": qty,
            "materials": [],
            "material_cost_per_unit": mat_price,
            "labor_price_per_unit": labor_price,
            "logistics_pct": 0,
            "markup_pct": 0,
            "risk_pct": 0,
            "total_cost_per_unit": round(mat_price + labor_price, 2),
            "final_price_per_unit": round(mat_price + labor_price, 2),
            "final_total": round((mat_price + labor_price) * qty, 2),
            "is_active": True,
            "original_qty": qty,
            "original_price": round(mat_price + labor_price, 2),
            "line_order": line_order,
        }
        lines.append(line)
        line_order += 1

    return {
        "lines": lines,
        "lines_count": len(lines),
        "skipped_count": skipped,
        "warnings": warnings,
        "detected_columns": mapping,
    }


# ── Export ─────────────────────────────────────────────────────────

def export_kss_to_excel(analysis: dict) -> bytes:
    """Generate Excel file from SMRAnalysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "КСС"

    header_font_w = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    inactive_font = Font(color="999999", strikethrough=True)
    summary_font = Font(bold=True, size=11)
    num_fmt = '#,##0.00'

    ws.merge_cells("A1:G1")
    ws["A1"] = analysis.get("name", "КСС")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Проект: {analysis.get('project_name', '')} | Версия: {analysis.get('version', 1)}"
    ws["A2"].font = Font(size=10, color="666666")

    headers = ["№", "Описание СМР", "Ед.", "Кол-во", "Материал/ед.", "Труд/ед.", "Общо"]
    col_widths = [6, 40, 8, 12, 14, 14, 16]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_num = 5
    for i, ln in enumerate(analysis.get("lines", [])):
        is_active = ln.get("is_active", True)
        font = inactive_font if not is_active else Font(size=10)
        ws.cell(row=row_num, column=1, value=i + 1).font = font
        ws.cell(row=row_num, column=2, value=ln.get("smr_type", "")).font = font
        ws.cell(row=row_num, column=3, value=ln.get("unit", "")).font = font
        for col, key in [(4, "qty"), (5, "material_cost_per_unit"), (6, "labor_price_per_unit"), (7, "final_total")]:
            cell = ws.cell(row=row_num, column=col, value=ln.get(key, 0))
            cell.font = font
            cell.number_format = num_fmt
            cell.border = thin_border
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1

    row_num += 1
    totals = analysis.get("totals", {})
    ws.cell(row=row_num, column=2, value="ОБЩО").font = summary_font
    ws.cell(row=row_num, column=7, value=totals.get("grand_total", 0)).font = summary_font
    ws.cell(row=row_num, column=7).number_format = num_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
