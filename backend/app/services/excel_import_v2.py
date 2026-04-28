"""
Service - Excel Import V2 (preview, templates, smart detection).
Additive layer on top of existing excel_import.py.
"""
import io
import uuid
from datetime import datetime, timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.db import db
from app.services.excel_import import COLUMN_KEYWORDS, parse_float


import unicodedata

def _normalize(s: str) -> str:
    """Normalize text for comparison: lowercase, strip, remove dots/dashes."""
    s = unicodedata.normalize("NFKD", s.lower().strip())
    s = s.replace(".", "").replace("-", "").replace("_", " ").replace("/", " ")
    return " ".join(s.split())  # collapse whitespace


def _match_header(val: str) -> str:
    """Match a header cell value to a known field."""
    v = _normalize(val)
    if not v:
        return ""
    for field, keywords in COLUMN_KEYWORDS.items():
        for kw in keywords:
            nkw = _normalize(kw)
            if nkw == v or nkw in v or v in nkw:
                return field
    return ""


def detect_excel_structure(file_bytes: bytes, sheet_name: str = None) -> dict:
    """Detect structure without importing."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb.active

    # Find header row
    best_row = 1
    best_score = 0
    for row_num in range(1, 8):
        score = 0
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            for val in row:
                if val and _match_header(str(val)):
                    score += 1
        if score > best_score:
            best_score = score
            best_row = row_num

    # Detect columns from header
    detected = {}
    headers_raw = {}
    for row in ws.iter_rows(min_row=best_row, max_row=best_row, values_only=False):
        for cell in row:
            v = str(cell.value or "").strip()
            if not v:
                continue
            col = get_column_letter(cell.column)
            headers_raw[col] = v
            field = _match_header(v)
            if field and field not in detected:
                detected[field] = col

    # Confidence
    expected = {"smr_type", "qty", "unit"}
    found = set(detected.keys())
    confidence = round(len(found & expected) / max(len(expected), 1), 2)
    if "smr_type" in found:
        confidence = max(confidence, 0.5)
    if "total_price" in found or "labor_price" in found:
        confidence = min(confidence + 0.2, 1.0)

    # Preview rows
    data_start = best_row + 1
    preview = []
    warnings = []
    total_rows = 0

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue
        total_rows += 1
        if len(preview) < 10:
            row_dict = {}
            for field, col in detected.items():
                idx = ord(col.upper()) - ord('A')
                row_dict[field] = row[idx] if idx < len(row) else None
            # Also include raw positional
            row_dict["_raw"] = [str(v) if v is not None else "" for v in row[:8]]
            preview.append(row_dict)

    if "smr_type" not in detected:
        warnings.append("Не е намерена колона за описание/дейност")
    if "qty" not in detected:
        warnings.append("Не е намерена колона за количество")

    return {
        "sheet_names": sheet_names,
        "selected_sheet": ws.title,
        "header_row_guess": best_row,
        "detected_columns": detected,
        "headers_raw": headers_raw,
        "preview_rows": preview,
        "total_rows": total_rows,
        "warnings": warnings,
        "confidence": confidence,
    }


def normalize_with_mapping(file_bytes: bytes, column_mapping: dict, sheet_name: str = None) -> dict:
    """Normalize rows using explicit column mapping."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Find header row (skip it)
    best_row = 1
    for row_num in range(1, 8):
        score = 0
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            for val in row:
                if val and _match_header(str(val)):
                    score += 1
        if score >= 2:
            best_row = row_num
            break

    data_start = best_row + 1
    rows = []
    warnings = []
    skipped = 0

    def _get(row_vals, field):
        col = column_mapping.get(field, "")
        if not col:
            return None
        idx = ord(col.upper()) - ord('A')
        return row_vals[idx] if idx < len(row_vals) else None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            skipped += 1
            continue
        smr = _get(row, "smr_type")
        if not smr:
            skipped += 1
            continue
        rows.append({
            "smr_type": str(smr).strip(),
            "unit": str(_get(row, "unit") or "m2").strip(),
            "qty": parse_float(_get(row, "qty")),
            "material_price": parse_float(_get(row, "material_price")),
            "labor_price": parse_float(_get(row, "labor_price")),
            "total_price": parse_float(_get(row, "total_price")),
        })

    return {"rows": rows, "warnings": warnings, "skipped": skipped, "total": len(rows)}


async def save_import_template(org_id: str, name: str, import_type: str, column_mapping: dict, created_by: str, **kwargs) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "org_id": org_id,
        "name": name,
        "import_type": import_type,
        "column_mapping": column_mapping,
        "sheet_name_default": kwargs.get("sheet_name_default"),
        "detected_headers": kwargs.get("detected_headers"),
        "notes": kwargs.get("notes"),
        "is_default": False,
        "created_by": created_by,
        "created_at": now,
        "updated_at": now,
    }
    await db.excel_import_templates.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


async def apply_template(org_id: str, template_id: str, file_bytes: bytes) -> dict:
    tpl = await db.excel_import_templates.find_one({"id": template_id, "org_id": org_id}, {"_id": 0})
    if not tpl:
        return {"error": "Template not found"}
    mapping = tpl.get("column_mapping", {})
    sheet = tpl.get("sheet_name_default")
    result = normalize_with_mapping(file_bytes, mapping, sheet)
    return {"template": tpl, **result}


# ═══════════════════════════════════════════════════════════════════
# CONSTRUCTION BUDGET IMPORT (Фаза 2.2)
# ═══════════════════════════════════════════════════════════════════

BUDGET_FIELDS = [
    "category", "smr_type", "unit", "qty", "material_price", "labor_price",
    "planned_man_hours", "akord", "planned_man_days", "total_per_unit",
    "total_price", "coefficient", "avg_daily_wage", "hours_per_day",
]

# Positional fallback for the specific construction budget template
BUDGET_POSITIONAL = {
    "category": 0, "smr_type": 1, "unit": 2, "qty": 3,
    "material_price": 4, "labor_price": 5, "planned_man_hours": 6,
    "akord": 7, "planned_man_days": 8, "total_per_unit": 9,
    "total_price": 10, "coefficient": 11, "avg_daily_wage": 12, "hours_per_day": 13,
}


def parse_construction_budget(file_bytes: bytes, sheet_name: str = None, column_mapping: dict = None) -> dict:
    """Parse a construction budget Excel template with man-hours formula fields."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Detect header row (look for more matches in rows 5-10 for this template)
    best_row = 7  # default for this template format
    best_score = 0
    for row_num in range(1, 12):
        score = 0
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            for val in row:
                if val and _match_header(str(val)):
                    score += 1
        if score > best_score:
            best_score = score
            best_row = row_num

    # Detect columns
    mapping = column_mapping or {}
    if not mapping:
        for row in ws.iter_rows(min_row=best_row, max_row=best_row, values_only=False):
            for cell in row:
                v = str(cell.value or "").lower().strip()
                if not v:
                    continue
                col = get_column_letter(cell.column)
                field = _match_header(v)
                if field and field not in mapping:
                    mapping[field] = col

    data_start = best_row + 1
    # Skip sub-header row if present
    for row in ws.iter_rows(min_row=data_start, max_row=data_start, values_only=True):
        all_vals = [str(v or "").strip().lower() for v in row]
        if any(v in ("ч.ч.", "акорд", "ч.д.", "") for v in all_vals[:5]):
            data_start += 1
        break

    lines = []
    warnings = []
    skipped = 0

    def _get(row_vals, field):
        if field in mapping:
            idx = ord(mapping[field].upper()) - ord('A')
            return row_vals[idx] if idx < len(row_vals) else None
        fb = BUDGET_POSITIONAL.get(field)
        if fb is not None and fb < len(row_vals):
            return row_vals[fb]
        return None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            skipped += 1
            continue

        activity = _get(row, "smr_type")
        if not activity:
            skipped += 1
            continue
        activity = str(activity).strip()

        category = str(_get(row, "category") or "").strip()
        unit = str(_get(row, "unit") or "m2").strip()
        qty = parse_float(_get(row, "qty"))
        mat_price = parse_float(_get(row, "material_price"))
        labor_price = parse_float(_get(row, "labor_price"))

        # Read formula fields (Mode A: full template)
        raw_man_hours = parse_float(_get(row, "planned_man_hours"))
        raw_akord = parse_float(_get(row, "akord"))
        raw_man_days = parse_float(_get(row, "planned_man_days"))
        raw_coeff = parse_float(_get(row, "coefficient")) or 2
        raw_wage = parse_float(_get(row, "avg_daily_wage"))
        raw_hpd = parse_float(_get(row, "hours_per_day")) or 8
        raw_total_pu = parse_float(_get(row, "total_per_unit"))
        raw_total = parse_float(_get(row, "total_price"))

        if qty <= 0:
            qty = 1
            warnings.append(f"{activity}: количество зададено на 1")

        # Calculate derived values
        labor_total = round(labor_price * qty, 2)
        materials_total = round(mat_price * qty, 2)
        coefficient = raw_coeff if raw_coeff > 0 else 2

        # Mode A: Excel has all values → use them directly
        # Mode B: Missing values → compute from budget_formula
        hours_per_day = raw_hpd if raw_hpd > 0 else 8

        if raw_akord > 0 and raw_man_days > 0 and raw_man_hours > 0:
            # Mode A: all values present in Excel
            akord = raw_akord
            man_days = raw_man_days
            man_hours = raw_man_hours
            mode = "A"
        else:
            # Mode B: compute from formula
            from app.services.budget_formula import calculate_budget_formula_sync
            wage_for_calc = raw_wage if raw_wage > 0 else 200  # default, will be overridden at commit
            r = calculate_budget_formula_sync(labor_total, coefficient, wage_for_calc, hours_per_day)
            akord = raw_akord if raw_akord > 0 else r["akord"]
            man_days = raw_man_days if raw_man_days > 0 else r["planned_man_days"]
            man_hours = raw_man_hours if raw_man_hours > 0 else r["planned_man_hours"]
            mode = "B"

        avg_wage = raw_wage if raw_wage > 0 else None  # None = compute from project team

        lines.append({
            "category": category,
            "activity_name": activity,
            "unit": unit,
            "qty": qty,
            "material_price": mat_price,
            "labor_price": labor_price,
            "labor_total": labor_total,
            "materials_total": materials_total,
            "coefficient": coefficient,
            "akord": akord,
            "planned_man_days": man_days,
            "planned_man_hours": man_hours,
            "avg_daily_wage": avg_wage,
            "hours_per_day": hours_per_day,
            "total_per_unit": raw_total_pu if raw_total_pu > 0 else round(mat_price + labor_price, 2),
            "total_price": raw_total if raw_total > 0 else round((mat_price + labor_price) * qty, 2),
            "import_mode": mode,
        })

    return {
        "lines": lines,
        "lines_count": len(lines),
        "skipped_count": skipped,
        "warnings": warnings,
        "detected_columns": mapping,
        "import_type": "construction_budget",
    }
