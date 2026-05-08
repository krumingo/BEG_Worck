"""
Routes - Offers / BOQ (M2) Endpoints.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
import uuid
import io
import logging

logger = logging.getLogger(__name__)

from app.db import db
from app.deps.auth import get_current_user, can_access_project, can_manage_project, get_user_project_ids
from app.deps.modules import require_m2
from app.utils.audit import log_audit
from ..models.offers import (
    OFFER_STATUSES, OFFER_UNITS,
    OfferCreate, OfferUpdate, OfferLinesUpdate, OfferReject,
    ActivityCatalogCreate, ActivityCatalogUpdate
)

router = APIRouter(tags=["Offers / BOQ"])

# ── Helpers ────────────────────────────────────────────────────────

async def get_next_offer_no(org_id: str) -> str:
    """Generate sequential offer number like OFF-0001"""
    last = await db.offers.find_one(
        {"org_id": org_id},
        {"_id": 0, "offer_no": 1},
        sort=[("created_at", -1)]
    )
    if last and last.get("offer_no"):
        try:
            num = int(last["offer_no"].split("-")[1]) + 1
        except (ValueError, TypeError) as e:
            logger.warning(f"offers.py parse error: {e}")
            num = 1
    else:
        num = 1
    return f"OFF-{num:04d}"


def compute_offer_line(line: dict) -> dict:
    """Compute line totals"""
    qty = line.get("qty", 0)
    material = line.get("material_unit_cost", 0)
    labor = line.get("labor_unit_cost", 0)
    line["line_material_cost"] = round(qty * material, 2)
    line["line_labor_cost"] = round(qty * labor, 2)
    line["line_total"] = round(line["line_material_cost"] + line["line_labor_cost"], 2)
    return line


def compute_offer_totals(offer: dict) -> dict:
    """Compute offer subtotal, vat, total"""
    lines = offer.get("lines", [])
    subtotal = sum(l.get("line_total", 0) for l in lines)
    vat_percent = offer.get("vat_percent", 0)
    vat_amount = round(subtotal * vat_percent / 100, 2)
    total = round(subtotal + vat_amount, 2)
    offer["subtotal"] = round(subtotal, 2)
    offer["vat_amount"] = vat_amount
    offer["total"] = total
    return offer


async def can_edit_offer(user: dict, offer: dict) -> bool:
    """Check if user can edit offer (must be Draft and have project access)"""
    if offer["status"] != "Draft":
        return False
    return await can_manage_project(user, offer["project_id"])


# ── Offer CRUD ─────────────────────────────────────────────────────

@router.post("/offers", status_code=201)
async def create_offer(data: OfferCreate, user: dict = Depends(require_m2)):
    if user["role"] not in ["Admin", "Owner", "SiteManager"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    project = await db.projects.find_one({"id": data.project_id, "org_id": user["org_id"]})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not await can_manage_project(user, data.project_id):
        raise HTTPException(status_code=403, detail="Not authorized for this project")
    
    now = datetime.now(timezone.utc).isoformat()
    offer_no = await get_next_offer_no(user["org_id"])
    
    lines = []
    for i, line in enumerate(data.lines):
        l = {
            "id": str(uuid.uuid4()),
            "activity_code": line.activity_code,
            "activity_name": line.activity_name,
            "unit": line.unit,
            "qty": line.qty,
            "material_unit_cost": line.material_unit_cost,
            "labor_unit_cost": line.labor_unit_cost,
            "labor_hours_per_unit": line.labor_hours_per_unit,
            "note": line.note,
            "sort_order": line.sort_order or i,
            "activity_type": line.activity_type or "Общо",
            "activity_subtype": line.activity_subtype or "",
        }
        lines.append(compute_offer_line(l))
    
    offer = {
        "id": str(uuid.uuid4()),
        "org_id": user["org_id"],
        "project_id": data.project_id,
        "offer_no": offer_no,
        "title": data.title,
        "status": "Draft",
        "version": 1,
        "parent_offer_id": None,
        "currency": data.currency,
        "vat_percent": data.vat_percent,
        "lines": lines,
        "notes": data.notes,
        "created_at": now,
        "updated_at": now,
        "sent_at": None,
        "accepted_at": None,
    }
    offer = compute_offer_totals(offer)
    
    await db.offers.insert_one(offer)
    await log_audit(user["org_id"], user["id"], user["email"], "offer_created", "offer", offer["id"], 
                    {"offer_no": offer_no, "title": data.title, "project_id": data.project_id})
    
    return {k: v for k, v in offer.items() if k != "_id"}


@router.get("/offers")
async def list_offers(
    user: dict = Depends(require_m2),
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
):
    query = {"org_id": user["org_id"]}
    
    if project_id:
        if not await can_access_project(user, project_id):
            raise HTTPException(status_code=403, detail="Access denied to project")
        query["project_id"] = project_id
    elif user["role"] not in ["Admin", "Owner", "Accountant"]:
        # Limit to assigned projects
        assigned = await get_user_project_ids(user["id"])
        query["project_id"] = {"$in": assigned}
    
    if status:
        query["status"] = status
    
    offers = await db.offers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    if search:
        s = search.lower()
        offers = [o for o in offers if s in o.get("offer_no", "").lower() or s in o.get("title", "").lower()]
    
    # Enrich with project info
    for o in offers:
        p = await db.projects.find_one({"id": o["project_id"]}, {"_id": 0, "code": 1, "name": 1})
        o["project_code"] = p["code"] if p else ""
        o["project_name"] = p["name"] if p else ""
        o["line_count"] = len(o.get("lines", []))
    
    return offers


@router.get("/offers/{offer_id}")
async def get_offer(offer_id: str, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]}, {"_id": 0})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if not await can_access_project(user, offer["project_id"]):
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Enrich
    p = await db.projects.find_one({"id": offer["project_id"]}, {"_id": 0, "code": 1, "name": 1})
    offer["project_code"] = p["code"] if p else ""
    offer["project_name"] = p["name"] if p else ""
    
    return offer


@router.put("/offers/{offer_id}")
async def update_offer(offer_id: str, data: OfferUpdate, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if not await can_edit_offer(user, offer):
        raise HTTPException(status_code=403, detail="Can only edit Draft offers you manage")
    
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.offers.update_one({"id": offer_id}, {"$set": update})
    
    # Recompute if vat changed
    if "vat_percent" in update:
        updated = await db.offers.find_one({"id": offer_id})
        updated = compute_offer_totals({k: v for k, v in updated.items() if k != "_id"})
        await db.offers.update_one({"id": offer_id}, {"$set": {
            "subtotal": updated["subtotal"],
            "vat_amount": updated["vat_amount"],
            "total": updated["total"],
        }})
    
    await log_audit(user["org_id"], user["id"], user["email"], "offer_updated", "offer", offer_id, update)
    return await db.offers.find_one({"id": offer_id}, {"_id": 0})


@router.put("/offers/{offer_id}/lines")
async def update_offer_lines(offer_id: str, data: OfferLinesUpdate, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if not await can_edit_offer(user, offer):
        raise HTTPException(status_code=403, detail="Can only edit Draft offers you manage")
    
    lines = []
    for i, line in enumerate(data.lines):
        l = {
            "id": str(uuid.uuid4()),
            "activity_code": line.activity_code,
            "activity_name": line.activity_name,
            "unit": line.unit,
            "qty": line.qty,
            "material_unit_cost": line.material_unit_cost,
            "labor_unit_cost": line.labor_unit_cost,
            "labor_hours_per_unit": line.labor_hours_per_unit,
            "note": line.note,
            "sort_order": line.sort_order or i,
            "activity_type": line.activity_type or "Общо",
            "activity_subtype": line.activity_subtype or "",
        }
        lines.append(compute_offer_line(l))
    
    now = datetime.now(timezone.utc).isoformat()
    updated = {
        "lines": lines,
        "updated_at": now,
    }
    
    # Compute totals
    offer["lines"] = lines
    offer["vat_percent"] = offer.get("vat_percent", 0)
    offer = compute_offer_totals(offer)
    updated["subtotal"] = offer["subtotal"]
    updated["vat_amount"] = offer["vat_amount"]
    updated["total"] = offer["total"]
    
    await db.offers.update_one({"id": offer_id}, {"$set": updated})
    await log_audit(user["org_id"], user["id"], user["email"], "offer_lines_updated", "offer", offer_id, 
                    {"line_count": len(lines)})
    
    return await db.offers.find_one({"id": offer_id}, {"_id": 0})


@router.post("/offers/{offer_id}/send")
async def send_offer(offer_id: str, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if not await can_manage_project(user, offer["project_id"]):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    if offer["status"] not in ["Draft"]:
        raise HTTPException(status_code=400, detail="Only Draft offers can be sent")
    if len(offer.get("lines", [])) == 0:
        raise HTTPException(status_code=400, detail="Offer must have at least one line")
    
    now = datetime.now(timezone.utc).isoformat()
    review_token = offer.get("review_token") or str(uuid.uuid4()).replace("-", "")[:24]
    
    # Save version snapshot before sending
    version_number = 1
    last_ver = await db.offer_versions.find_one(
        {"org_id": user["org_id"], "offer_id": offer_id}, sort=[("version_number", -1)]
    )
    if last_ver:
        version_number = last_ver["version_number"] + 1
    
    snapshot = {
        "offer_no": offer.get("offer_no"), "title": offer.get("title"),
        "status": "Sent", "version": offer.get("version"),
        "currency": offer.get("currency"), "vat_percent": offer.get("vat_percent"),
        "notes": offer.get("notes"), "lines": offer.get("lines", []),
        "subtotal": offer.get("subtotal"), "vat_amount": offer.get("vat_amount"),
        "total": offer.get("total"),
    }
    await db.offer_versions.insert_one({
        "id": str(uuid.uuid4()), "org_id": user["org_id"],
        "project_id": offer["project_id"], "offer_id": offer_id,
        "version_number": version_number, "created_at": now,
        "created_by": user["id"], "note": "Автоматичен snapshot при изпращане",
        "snapshot_json": snapshot, "is_auto_backup": True,
    })
    
    await db.offers.update_one({"id": offer_id}, {"$set": {
        "status": "Sent", "sent_at": now, "sent_by": user["id"],
        "review_token": review_token, "updated_at": now,
    }})
    
    # Record event
    await db.offer_events.insert_one({
        "id": str(uuid.uuid4()), "org_id": user["org_id"],
        "offer_id": offer_id, "event_type": "sent",
        "actor": user["email"], "created_at": now,
        "details": {"sent_by": user["id"]},
    })
    
    await log_audit(user["org_id"], user["id"], user["email"], "offer_sent", "offer", offer_id, 
                    {"offer_no": offer["offer_no"]})
    
    result = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    result["review_url"] = f"/offers/review/{review_token}"
    return result


@router.post("/offers/{offer_id}/accept")
async def accept_offer(offer_id: str, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if user["role"] not in ["Admin", "Owner"]:
        raise HTTPException(status_code=403, detail="Only Admin/Owner can accept offers")
    if offer["status"] != "Sent":
        raise HTTPException(status_code=400, detail="Only Sent offers can be accepted")
    
    now = datetime.now(timezone.utc).isoformat()
    await db.offers.update_one({"id": offer_id}, {"$set": {
        "status": "Accepted",
        "accepted_at": now,
        "updated_at": now,
    }})
    await log_audit(user["org_id"], user["id"], user["email"], "offer_accepted", "offer", offer_id,
                    {"offer_no": offer["offer_no"], "total": offer.get("total", 0)})
    
    return await db.offers.find_one({"id": offer_id}, {"_id": 0})


@router.post("/offers/{offer_id}/reject")
async def reject_offer(offer_id: str, data: OfferReject, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if user["role"] not in ["Admin", "Owner"]:
        raise HTTPException(status_code=403, detail="Only Admin/Owner can reject offers")
    if offer["status"] != "Sent":
        raise HTTPException(status_code=400, detail="Only Sent offers can be rejected")
    
    now = datetime.now(timezone.utc).isoformat()
    await db.offers.update_one({"id": offer_id}, {"$set": {
        "status": "Rejected",
        "reject_reason": data.reason,
        "updated_at": now,
    }})
    await log_audit(user["org_id"], user["id"], user["email"], "offer_rejected", "offer", offer_id,
                    {"offer_no": offer["offer_no"], "reason": data.reason})
    
    return await db.offers.find_one({"id": offer_id}, {"_id": 0})


@router.post("/offers/{offer_id}/new-version")
async def create_offer_version(offer_id: str, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if not await can_manage_project(user, offer["project_id"]):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    if offer["status"] not in ["Sent", "Accepted", "Rejected"]:
        raise HTTPException(status_code=400, detail="Can only version non-Draft offers")
    
    now = datetime.now(timezone.utc).isoformat()
    new_version = offer.get("version", 1) + 1
    
    # Clone lines with new IDs
    new_lines = []
    for line in offer.get("lines", []):
        l = {k: v for k, v in line.items() if k != "_id"}
        l["id"] = str(uuid.uuid4())
        new_lines.append(l)
    
    new_offer = {
        "id": str(uuid.uuid4()),
        "org_id": user["org_id"],
        "project_id": offer["project_id"],
        "offer_no": offer["offer_no"],  # Keep same offer_no
        "title": offer["title"],
        "status": "Draft",
        "version": new_version,
        "parent_offer_id": offer["id"],
        "currency": offer["currency"],
        "vat_percent": offer["vat_percent"],
        "lines": new_lines,
        "subtotal": offer["subtotal"],
        "vat_amount": offer["vat_amount"],
        "total": offer["total"],
        "notes": offer.get("notes", ""),
        "created_at": now,
        "updated_at": now,
        "sent_at": None,
        "accepted_at": None,
    }
    
    await db.offers.insert_one(new_offer)
    await log_audit(user["org_id"], user["id"], user["email"], "offer_versioned", "offer", new_offer["id"],
                    {"offer_no": offer["offer_no"], "version": new_version, "from_version": offer.get("version", 1)})
    
    return {k: v for k, v in new_offer.items() if k != "_id"}


@router.delete("/offers/{offer_id}")
async def delete_offer(offer_id: str, user: dict = Depends(require_m2)):
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    if user["role"] not in ["Admin", "Owner"]:
        raise HTTPException(status_code=403, detail="Only Admin/Owner can delete offers")
    if offer["status"] == "Accepted":
        raise HTTPException(status_code=400, detail="Cannot delete accepted offers")
    
    await db.offers.delete_one({"id": offer_id})
    await log_audit(user["org_id"], user["id"], user["email"], "offer_deleted", "offer", offer_id,
                    {"offer_no": offer["offer_no"]})
    return {"ok": True}


# ── Activity Catalog CRUD ──────────────────────────────────────────

@router.get("/activity-catalog")
async def list_activity_catalog(
    user: dict = Depends(require_m2),
    project_id: Optional[str] = None,
    active_only: bool = True,
):
    query = {"org_id": user["org_id"]}
    if project_id:
        if not await can_access_project(user, project_id):
            raise HTTPException(status_code=403, detail="Access denied")
        query["project_id"] = project_id
    if active_only:
        query["active"] = True
    
    items = await db.activity_catalog.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return items


@router.post("/activity-catalog", status_code=201)
async def create_activity(data: ActivityCatalogCreate, user: dict = Depends(require_m2)):
    if user["role"] not in ["Admin", "Owner", "SiteManager"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    project = await db.projects.find_one({"id": data.project_id, "org_id": user["org_id"]})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not await can_manage_project(user, data.project_id):
        raise HTTPException(status_code=403, detail="Not authorized for this project")
    
    now = datetime.now(timezone.utc).isoformat()
    item = {
        "id": str(uuid.uuid4()),
        "org_id": user["org_id"],
        "project_id": data.project_id,
        "code": data.code,
        "name": data.name,
        "default_unit": data.default_unit,
        "default_material_unit_cost": data.default_material_unit_cost,
        "default_labor_unit_cost": data.default_labor_unit_cost,
        "default_labor_hours_per_unit": data.default_labor_hours_per_unit,
        "active": data.active,
        "created_at": now,
        "updated_at": now,
    }
    await db.activity_catalog.insert_one(item)
    return {k: v for k, v in item.items() if k != "_id"}


@router.put("/activity-catalog/{item_id}")
async def update_activity(item_id: str, data: ActivityCatalogUpdate, user: dict = Depends(require_m2)):
    item = await db.activity_catalog.find_one({"id": item_id, "org_id": user["org_id"]})
    if not item:
        raise HTTPException(status_code=404, detail="Activity not found")
    if not await can_manage_project(user, item["project_id"]):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.activity_catalog.update_one({"id": item_id}, {"$set": update})
    return await db.activity_catalog.find_one({"id": item_id}, {"_id": 0})


@router.delete("/activity-catalog/{item_id}")
async def delete_activity(item_id: str, user: dict = Depends(require_m2)):
    item = await db.activity_catalog.find_one({"id": item_id, "org_id": user["org_id"]})
    if not item:
        raise HTTPException(status_code=404, detail="Activity not found")
    if not await can_manage_project(user, item["project_id"]):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    await db.activity_catalog.delete_one({"id": item_id})
    return {"ok": True}


@router.get("/offer-enums")
async def get_offer_enums():
    return {
        "statuses": OFFER_STATUSES,
        "units": OFFER_UNITS,
    }


# ── Offer Export (PDF + XLSX) ──────────────────────────────────────

UNIT_LABELS_BG = {"m2": "м2", "m": "м", "pcs": "бр", "hours": "часа", "lot": "к-т", "kg": "кг", "l": "л"}
STATUS_LABELS_BG = {"Draft": "Чернова", "Sent": "Изпратена", "Accepted": "Одобрена", "Rejected": "Отказана", "NeedsRevision": "Корекция", "Archived": "Архивирана"}

@router.get("/offers/{offer_id}/pdf")
async def export_offer_pdf(offer_id: str, user: dict = Depends(require_m2)):
    """Export offer as PDF"""
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]}, {"_id": 0})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    
    org = await db.organizations.find_one({"id": user["org_id"]}, {"_id": 0})
    project = await db.projects.find_one({"id": offer.get("project_id")}, {"_id": 0, "code": 1, "name": 1, "address_text": 1}) if offer.get("project_id") else None
    
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")
    
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuBold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        fn, fb = 'DejaVu', 'DejaVuBold'
    except Exception:
        fn, fb = 'Helvetica', 'Helvetica-Bold'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    sTitle = ParagraphStyle('T', parent=styles['Title'], fontName=fb, fontSize=16, textColor=colors.HexColor('#333'))
    sH2 = ParagraphStyle('H', parent=styles['Heading2'], fontName=fb, fontSize=11, textColor=colors.HexColor('#555'))
    sN = ParagraphStyle('N', parent=styles['Normal'], fontName=fn, fontSize=9, leading=13)
    sS = ParagraphStyle('S', parent=styles['Normal'], fontName=fn, fontSize=8, textColor=colors.HexColor('#888'))
    sB = ParagraphStyle('B', parent=styles['Normal'], fontName=fb, fontSize=9, leading=13)
    
    is_extra = offer.get("offer_type") == "extra"
    label = "Допълнителна оферта" if is_extra else "Оферта"
    org_name = org.get("name", "") if org else ""
    
    elements = []
    
    # Header
    hdr = [[Paragraph(f"<b>{org_name}</b>", sB), Paragraph(f"<b>{label}</b>", ParagraphStyle('R', parent=sTitle, alignment=2, fontSize=13))],
           [Paragraph(org.get("address", "") if org else "", sS), Paragraph(f"<b>№ {offer.get('offer_no', '')}</b>  v{offer.get('version', 1)}", ParagraphStyle('R', parent=sB, alignment=2, fontSize=13))]]
    ht = RLTable(hdr, colWidths=[300, 230])
    ht.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2)]))
    elements.extend([ht, Spacer(1, 5*mm)])
    
    # Meta
    meta = [f"Дата: {offer.get('created_at', '')[:10]}", f"Статус: {STATUS_LABELS_BG.get(offer.get('status', ''), offer.get('status', ''))}"]
    if project:
        meta.append(f"Обект: {project.get('code', '')} - {project.get('name', '')}")
    elements.extend([Paragraph(" | ".join(meta), sS), Spacer(1, 5*mm)])
    
    if offer.get("title"):
        elements.extend([Paragraph(f"<b>{offer['title']}</b>", sB), Spacer(1, 3*mm)])
    
    # Lines table
    cw = [22, 180, 35, 40, 55, 55, 55, 55]
    line_data = [["№", "Описание", "Ед.", "К-во", "Мат./ед", "Труд/ед", "Мат. общо", "Общо"]]
    for i, l in enumerate(offer.get("lines", [])):
        qty = l.get("qty", 0)
        mu = l.get("material_unit_cost", 0)
        lu = l.get("labor_unit_cost", 0)
        lmc = l.get("line_material_cost", round(qty * mu, 2))
        lt = l.get("line_total", round(qty * (mu + lu), 2))
        note = ""
        if l.get("note"):
            note = f"\n{l['note']}"
        line_data.append([str(i+1), Paragraph(f"{l.get('activity_name', '')}{note}", sN), UNIT_LABELS_BG.get(l.get("unit", ""), l.get("unit", "")), f"{qty:.1f}", f"{mu:.2f}", f"{lu:.2f}", f"{lmc:.2f}", f"{lt:.2f}"])
    
    lt = RLTable(line_data, colWidths=cw, repeatRows=1)
    lt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0,0), (-1,0), fb), ('FONTNAME', (0,1), (-1,-1), fn),
        ('FONTSIZE', (0,0), (-1,-1), 8), ('ALIGN', (2,0), (-1,-1), 'RIGHT'), ('ALIGN', (0,0), (1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#ccc')),
        ('TOPPADDING', (0,0), (-1,-1), 3), ('BOTTOMPADDING', (0,0), (-1,-1), 3), ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elements.extend([lt, Spacer(1, 5*mm)])
    
    # Totals
    curr = offer.get("currency", "BGN")
    sub = offer.get("subtotal", 0)
    vp = offer.get("vat_percent", 0)
    va = offer.get("vat_amount", 0)
    tot = offer.get("total", 0)
    td = [["", "", "", "", "", "", "Междинна сума:", f"{sub:.2f} {curr}"],
          ["", "", "", "", "", "", f"ДДС ({vp}%):", f"{va:.2f} {curr}"],
          ["", "", "", "", "", "", "ОБЩО:", f"{tot:.2f} {curr}"]]
    tt = RLTable(td, colWidths=cw)
    tt.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), fn), ('FONTNAME', (6,2), (6,2), fb),
        ('FONTSIZE', (0,0), (-1,-1), 9), ('FONTSIZE', (6,2), (7,2), 11),
        ('ALIGN', (6,0), (-1,-1), 'RIGHT'), ('LINEABOVE', (6,2), (-1,2), 1, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2)]))
    elements.append(tt)
    
    if offer.get("notes"):
        elements.extend([Spacer(1, 6*mm), Paragraph("Бележки:", sH2), Paragraph(offer["notes"], sN)])
    
    doc.build(elements)
    buffer.seek(0)
    fn_name = f"offer_{offer.get('offer_no', offer_id)}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
                            headers={"Content-Disposition": f'attachment; filename="{fn_name}"'})


@router.get("/offers/{offer_id}/xlsx")
async def export_offer_xlsx(offer_id: str, user: dict = Depends(require_m2)):
    """Export offer as XLSX"""
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]}, {"_id": 0})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    
    org = await db.organizations.find_one({"id": user["org_id"]}, {"_id": 0})
    project = await db.projects.find_one({"id": offer.get("project_id")}, {"_id": 0, "code": 1, "name": 1}) if offer.get("project_id") else None
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оферта"
    
    thin = Side(style='thin', color='CCCCCC')
    header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    bold_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=9)
    normal_font = Font(size=9)
    
    # Header info
    ws.merge_cells('A1:H1')
    ws['A1'] = org.get("name", "") if org else ""
    ws['A1'].font = Font(bold=True, size=14)
    
    is_extra = offer.get("offer_type") == "extra"
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{'Допълнителна оферта' if is_extra else 'Оферта'} {offer.get('offer_no', '')} v{offer.get('version', 1)}"
    ws['A2'].font = bold_font
    
    ws['A3'] = "Проект:"
    ws['B3'] = f"{project.get('code', '')} - {project.get('name', '')}" if project else ""
    ws['A4'] = "Дата:"
    ws['B4'] = offer.get("created_at", "")[:10]
    ws['C4'] = "Статус:"
    ws['D4'] = STATUS_LABELS_BG.get(offer.get("status", ""), offer.get("status", ""))
    ws['E4'] = "Валута:"
    ws['F4'] = offer.get("currency", "BGN")
    
    # Column headers
    row = 6
    headers = ["№", "Описание / Вид СМР", "Мярка", "Количество", "Цена материал", "Цена труд", "Материал общо", "Общо"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        c.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    
    # Lines
    for i, l in enumerate(offer.get("lines", [])):
        row += 1
        qty = l.get("qty", 0)
        mu = l.get("material_unit_cost", 0)
        lu = l.get("labor_unit_cost", 0)
        lmc = round(qty * mu, 2)
        lt = round(qty * (mu + lu), 2)
        
        desc = l.get("activity_name", "")
        if l.get("note"):
            desc += f" ({l['note']})"
        
        vals = [i+1, desc, UNIT_LABELS_BG.get(l.get("unit", ""), l.get("unit", "")), qty, mu, lu, lmc, lt]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = normal_font
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if col >= 4:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
    
    # Totals
    row += 2
    ws.cell(row=row, column=7, value="Междинна сума:").font = bold_font
    ws.cell(row=row, column=8, value=offer.get("subtotal", 0)).font = bold_font
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=7, value=f"ДДС ({offer.get('vat_percent', 0)}%):").font = normal_font
    ws.cell(row=row, column=8, value=offer.get("vat_amount", 0)).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=7, value="ОБЩО:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=8, value=offer.get("total", 0)).font = Font(bold=True, size=12)
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    
    if offer.get("notes"):
        row += 2
        ws.cell(row=row, column=1, value="Бележки:").font = bold_font
        ws.cell(row=row+1, column=1, value=offer["notes"]).font = normal_font
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fn_name = f"offer_{offer.get('offer_no', offer_id)}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f'attachment; filename="{fn_name}"'})


# ── Offer Import from XLSX ─────────────────────────────────────────

@router.post("/offers/import-preview")
async def import_offer_preview(file: UploadFile = File(...), user: dict = Depends(require_m2)):
    """Parse uploaded XLSX and return preview of lines for review"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Поддържат се само .xlsx файлове")
    
    import openpyxl
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Файлът е празен")
    
    # Find header row (look for known column names)
    header_row = None
    header_idx = 0
    known_headers = {"описание", "вид смр", "мярка", "количество", "цена", "материал", "труд", "общо", "description", "unit", "qty", "price"}
    
    for i, row in enumerate(rows[:10]):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        matches = sum(1 for h in row_lower if any(k in h for k in known_headers))
        if matches >= 2:
            header_row = row
            header_idx = i
            break
    
    if not header_row:
        # Try first row as header
        header_row = rows[0]
        header_idx = 0
    
    # Map columns
    col_map = {}
    UNIT_REVERSE = {"м2": "m2", "м": "m", "бр": "pcs", "часа": "hours", "к-т": "lot", "кг": "kg", "л": "l"}
    
    for ci, cell in enumerate(header_row):
        h = str(cell).lower().strip() if cell else ""
        if any(k in h for k in ["описание", "вид смр", "description", "наименование", "дейност"]):
            col_map["description"] = ci
        elif any(k in h for k in ["мярка", "ед.", "unit"]):
            col_map["unit"] = ci
        elif any(k in h for k in ["количество", "к-во", "qty"]):
            col_map["qty"] = ci
        elif any(k in h for k in ["цена материал", "мат. цена", "material"]):
            col_map["material_price"] = ci
        elif any(k in h for k in ["цена труд", "труд", "labor"]):
            col_map["labor_price"] = ci
        elif any(k in h for k in ["обща цена", "общо", "total", "стойност"]):
            col_map["total"] = ci
        elif any(k in h for k in ["бележк", "note"]):
            col_map["notes"] = ci
        elif any(k in h for k in ["етаж", "floor"]):
            col_map["floor"] = ci
        elif any(k in h for k in ["помещение", "room"]):
            col_map["room"] = ci
        elif any(k in h for k in ["зона", "zone"]):
            col_map["zone"] = ci
    
    # Parse data rows
    lines = []
    warnings = []
    
    for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        
        desc = str(row[col_map["description"]]).strip() if "description" in col_map and col_map["description"] < len(row) and row[col_map["description"]] else ""
        if not desc:
            continue
        
        unit_raw = str(row[col_map["unit"]]).strip() if "unit" in col_map and col_map["unit"] < len(row) and row[col_map["unit"]] else "pcs"
        unit = UNIT_REVERSE.get(unit_raw.lower(), unit_raw.lower()) if unit_raw else "pcs"
        if unit not in OFFER_UNITS:
            warnings.append(f"Ред {ri}: неразпозната мярка '{unit_raw}'")
            unit = "pcs"
        
        def safe_float(val):
            try: return float(val) if val else 0
            except (ValueError, TypeError): return 0
        
        qty = safe_float(row[col_map["qty"]] if "qty" in col_map and col_map["qty"] < len(row) else 0)
        mat_price = safe_float(row[col_map["material_price"]] if "material_price" in col_map and col_map["material_price"] < len(row) else 0)
        lab_price = safe_float(row[col_map["labor_price"]] if "labor_price" in col_map and col_map["labor_price"] < len(row) else 0)
        total = safe_float(row[col_map["total"]] if "total" in col_map and col_map["total"] < len(row) else 0)
        
        # If only total given, split 50/50
        if total > 0 and mat_price == 0 and lab_price == 0 and qty > 0:
            price_per_unit = total / qty
            mat_price = round(price_per_unit * 0.4, 2)
            lab_price = round(price_per_unit * 0.6, 2)
        
        if qty == 0:
            warnings.append(f"Ред {ri}: количество е 0 за '{desc}'")
        if mat_price == 0 and lab_price == 0:
            warnings.append(f"Ред {ri}: няма цена за '{desc}'")
        
        note = str(row[col_map["notes"]]).strip() if "notes" in col_map and col_map["notes"] < len(row) and row[col_map["notes"]] else ""
        floor = str(row[col_map["floor"]]).strip() if "floor" in col_map and col_map["floor"] < len(row) and row[col_map["floor"]] else ""
        room = str(row[col_map["room"]]).strip() if "room" in col_map and col_map["room"] < len(row) and row[col_map["room"]] else ""
        zone = str(row[col_map["zone"]]).strip() if "zone" in col_map and col_map["zone"] < len(row) and row[col_map["zone"]] else ""
        
        location_parts = [p for p in [floor and f"Ет.{floor}", room, zone] if p]
        if location_parts and note:
            note = f"Локация: {', '.join(location_parts)}; {note}"
        elif location_parts:
            note = f"Локация: {', '.join(location_parts)}"
        
        lines.append({
            "row_number": ri,
            "description": desc,
            "unit": unit,
            "qty": qty,
            "material_price": mat_price,
            "labor_price": lab_price,
            "total": total,
            "note": note,
        })
    
    return {
        "file_name": file.filename,
        "total_rows": len(rows),
        "parsed_lines": len(lines),
        "column_mapping": col_map,
        "headers_found": [str(c) for c in header_row if c],
        "warnings": warnings,
        "lines": lines,
    }


@router.post("/offers/import-confirm", status_code=201)
async def import_offer_confirm(data: dict, user: dict = Depends(require_m2)):
    """Create offer from imported/previewed lines"""
    if user["role"] not in ["Admin", "Owner", "SiteManager"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    project_id = data.get("project_id")
    if not project_id:
        raise HTTPException(status_code=400, detail="project_id required")
    
    project = await db.projects.find_one({"id": project_id, "org_id": user["org_id"]})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    import_lines = data.get("lines", [])
    if not import_lines:
        raise HTTPException(status_code=400, detail="No lines to import")
    
    offer_type = data.get("offer_type", "main")
    title = data.get("title", f"Импортирана оферта ({datetime.now().strftime('%d.%m.%Y')})")
    currency = data.get("currency", "BGN")
    vat_percent = float(data.get("vat_percent", 20))
    
    now = datetime.now(timezone.utc).isoformat()
    offer_no = await get_next_offer_no(user["org_id"])
    
    lines = []
    for i, il in enumerate(import_lines):
        qty = float(il.get("qty", 1))
        mu = float(il.get("material_price", 0))
        lu = float(il.get("labor_price", 0))
        l = {
            "id": str(uuid.uuid4()),
            "activity_code": None,
            "activity_name": il.get("description", ""),
            "unit": il.get("unit", "pcs"),
            "qty": qty,
            "material_unit_cost": mu,
            "labor_unit_cost": lu,
            "labor_hours_per_unit": None,
            "line_material_cost": round(qty * mu, 2),
            "line_labor_cost": round(qty * lu, 2),
            "line_total": round(qty * (mu + lu), 2),
            "note": il.get("note", ""),
            "sort_order": i,
            "activity_type": "Общо",
            "activity_subtype": "",
        }
        lines.append(l)
    
    subtotal = sum(l["line_total"] for l in lines)
    vat_amount = round(subtotal * vat_percent / 100, 2)
    total = round(subtotal + vat_amount, 2)
    
    offer = {
        "id": str(uuid.uuid4()),
        "org_id": user["org_id"],
        "project_id": project_id,
        "offer_no": offer_no,
        "title": title,
        "offer_type": offer_type,
        "status": "Draft",
        "version": 1,
        "parent_offer_id": None,
        "currency": currency,
        "vat_percent": vat_percent,
        "lines": lines,
        "subtotal": round(subtotal, 2),
        "vat_amount": vat_amount,
        "total": total,
        "notes": data.get("notes", f"Импортирана от файл: {data.get('file_name', '')}"),
        "created_at": now,
        "updated_at": now,
        "sent_at": None,
        "accepted_at": None,
        "import_source": data.get("file_name"),
    }
    
    await db.offers.insert_one(offer)
    await log_audit(user["org_id"], user["id"], user["email"], "offer_imported", "offer", offer["id"],
                    {"offer_no": offer_no, "lines": len(lines), "source": data.get("file_name")})
    
    return {k: v for k, v in offer.items() if k != "_id"}


# ── Import Template Download ───────────────────────────────────────

@router.get("/offer-import-template")
async def download_import_template(user: dict = Depends(require_m2)):
    """Download XLSX template for offer import"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оферта"
    
    thin = Side(style='thin', color='CCCCCC')
    hdr_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    hdr_font = Font(bold=True, size=10)
    
    headers = ["Описание / Вид СМР", "Мярка", "Количество", "Цена материал", "Цена труд", "Обща цена", "Бележки", "Етаж", "Помещение", "Зона"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    # Example rows
    examples = [
        ["Направа на мазилка по стени", "м2", 25, 8.50, 12.00, 512.50, "", "1", "Спалня", "Южна стена"],
        ["Боядисване на стени и тавани", "м2", 40, 4.50, 6.00, 420.00, "2 слоя латекс", "1", "Хол", ""],
        ["Монтаж на ел. ключове/контакти", "бр", 8, 35.00, 25.00, 480.00, "Schneider серия", "", "", ""],
    ]
    for ri, ex in enumerate(examples, 2):
        for ci, v in enumerate(ex, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(size=9, italic=True, color='888888')
            if ci >= 3 and isinstance(v, (int, float)):
                c.number_format = '#,##0.00'
    
    # Instructions sheet
    ws2 = wb.create_sheet("Инструкции")
    instructions = [
        "ИНСТРУКЦИИ ЗА ИМПОРТ НА ОФЕРТИ",
        "",
        "Задължителни колони:",
        "  - Описание / Вид СМР (текст)",
        "  - Мярка: м2, м, бр, часа, к-т, кг, л",
        "  - Количество (число)",
        "",
        "Поне едно от:",
        "  - Цена материал + Цена труд (числа)",
        "  - ИЛИ Обща цена (число, ще се разпредели 40/60)",
        "",
        "Допълнителни колони (незадължителни):",
        "  - Бележки",
        "  - Етаж, Помещение, Зона (за локация)",
        "",
        "Поддържани мерни единици:",
        "  м2, м, бр, часа, к-т, кг, л",
        "",
        "Формат: .xlsx (Excel 2007+)",
    ]
    for ri, txt in enumerate(instructions, 1):
        ws2.cell(row=ri, column=1, value=txt).font = Font(size=10, bold=(ri == 1))
    ws2.column_dimensions['A'].width = 60
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": 'attachment; filename="BEG_Work_Offer_Import_Template.xlsx"'})

@router.get("/offers/review/{review_token}")
async def get_offer_review(review_token: str):
    """Public endpoint - get offer for client review (no auth required)"""
    offer = await db.offers.find_one({"review_token": review_token}, {"_id": 0})
    if not offer:
        raise HTTPException(status_code=404, detail="Офертата не е намерена")
    
    # Record view event (only first time)
    if offer["status"] == "Sent":
        now = datetime.now(timezone.utc).isoformat()
        await db.offers.update_one({"review_token": review_token}, {"$set": {
            "viewed_at": now, "updated_at": now,
        }})
        existing_view = await db.offer_events.find_one({
            "offer_id": offer["id"], "event_type": "viewed",
        })
        if not existing_view:
            await db.offer_events.insert_one({
                "id": str(uuid.uuid4()), "org_id": offer["org_id"],
                "offer_id": offer["id"], "event_type": "viewed",
                "actor": "client", "created_at": now, "details": {},
            })
    
    # Get project info
    project = await db.projects.find_one({"id": offer["project_id"]}, {"_id": 0, "code": 1, "name": 1, "address_text": 1})
    
    # Get org info
    org = await db.organizations.find_one({"id": offer["org_id"]}, {"_id": 0, "name": 1, "phone": 1, "email": 1})
    
    return {
        "offer_no": offer.get("offer_no"),
        "title": offer.get("title"),
        "status": offer.get("status"),
        "version": offer.get("version", 1),
        "offer_type": offer.get("offer_type"),
        "currency": offer.get("currency"),
        "vat_percent": offer.get("vat_percent"),
        "lines": offer.get("lines", []),
        "subtotal": offer.get("subtotal"),
        "vat_amount": offer.get("vat_amount"),
        "total": offer.get("total"),
        "notes": offer.get("notes"),
        "sent_at": offer.get("sent_at"),
        "project_code": project.get("code") if project else "",
        "project_name": project.get("name") if project else "",
        "project_address": project.get("address_text") if project else "",
        "company_name": org.get("name") if org else "",
        "company_phone": org.get("phone") if org else "",
        "company_email": org.get("email") if org else "",
    }


@router.post("/offers/review/{review_token}/respond")
async def respond_to_offer(review_token: str, data: dict):
    """Public endpoint - client responds to offer (no auth required)"""
    offer = await db.offers.find_one({"review_token": review_token})
    if not offer:
        raise HTTPException(status_code=404, detail="Офертата не е намерена")
    if offer["status"] not in ["Sent"]:
        raise HTTPException(status_code=400, detail="Тази оферта вече е обработена")
    
    action = data.get("action")  # approve / reject / revision
    comment = data.get("comment", "")
    client_name = data.get("client_name", "")
    now = datetime.now(timezone.utc).isoformat()
    
    if action == "approve":
        new_status = "Accepted"
        event_type = "approved_by_client"
        update_fields = {"status": new_status, "accepted_at": now, "client_comment": comment, "client_name": client_name, "updated_at": now}
    elif action == "reject":
        new_status = "Rejected"
        event_type = "rejected_by_client"
        update_fields = {"status": new_status, "reject_reason": comment, "client_comment": comment, "client_name": client_name, "updated_at": now}
    elif action == "revision":
        new_status = "NeedsRevision"
        event_type = "revision_requested"
        update_fields = {"status": new_status, "revision_comment": comment, "client_comment": comment, "client_name": client_name, "updated_at": now}
    else:
        raise HTTPException(status_code=400, detail="Невалидно действие")
    
    await db.offers.update_one({"review_token": review_token}, {"$set": update_fields})
    
    await db.offer_events.insert_one({
        "id": str(uuid.uuid4()), "org_id": offer["org_id"],
        "offer_id": offer["id"], "event_type": event_type,
        "actor": client_name or "client", "created_at": now,
        "details": {"comment": comment, "client_name": client_name},
    })
    
    return {"ok": True, "status": new_status}


# ── Offer Events History ───────────────────────────────────────────

@router.get("/offers/{offer_id}/events")
async def get_offer_events(offer_id: str, user: dict = Depends(require_m2)):
    """Get event history for an offer"""
    offer = await db.offers.find_one({"id": offer_id, "org_id": user["org_id"]})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    
    events = await db.offer_events.find(
        {"offer_id": offer_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return events
