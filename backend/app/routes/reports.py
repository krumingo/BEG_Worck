"""
Routes - Reports (Prices, Turnover, Company Finance, etc.)
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from typing import Optional
from datetime import datetime, timezone, timedelta
from calendar import monthrange
import re

from app.db import db
from app.deps.auth import get_current_user
from app.deps.modules import require_m5

router = APIRouter(tags=["Reports"])


# ── Prices (Purchase History) ─────────────────────────────────────

@router.get("/prices")
async def get_price_history(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str = Query("created_at"),
    sort_dir: str = Query("desc"),
    search: Optional[str] = None,
    item_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    project_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """
    Get purchase price history from invoice lines.
    Shows item prices from different suppliers over time.
    """
    # Build pipeline
    pipeline = [
        {"$match": {"org_id": user["org_id"]}},
    ]
    
    # Filter by allocation
    if project_id:
        pipeline.append({"$match": {"allocations": {"$elemMatch": {"type": "project", "ref_id": project_id}}}})
    if warehouse_id:
        pipeline.append({"$match": {"allocations": {"$elemMatch": {"type": "warehouse", "ref_id": warehouse_id}}}})
    
    # Lookup invoice for supplier and date
    pipeline.append({
        "$lookup": {
            "from": "invoices",
            "localField": "invoice_id",
            "foreignField": "id",
            "as": "invoice"
        }
    })
    pipeline.append({"$unwind": {"path": "$invoice", "preserveNullAndEmptyArrays": True}})
    
    # Filter by supplier
    if supplier_id:
        pipeline.append({"$match": {"invoice.supplier_counterparty_id": supplier_id}})
    
    # Filter by date
    if date_from:
        pipeline.append({"$match": {"invoice.issue_date": {"$gte": date_from}}})
    if date_to:
        pipeline.append({"$match": {"invoice.issue_date": {"$lte": date_to}}})
    
    # Search by description
    if search:
        pipeline.append({"$match": {"description": {"$regex": search, "$options": "i"}}})
    
    # Project fields
    pipeline.append({
        "$project": {
            "_id": 0,
            "line_id": "$id",
            "invoice_id": 1,
            "invoice_no": "$invoice.invoice_no",
            "invoice_date": "$invoice.issue_date",
            "supplier_id": "$invoice.supplier_counterparty_id",
            "description": 1,
            "unit": 1,
            "qty": 1,
            "unit_price": 1,
            "line_total_ex_vat": 1,
            "vat_amount": 1,
            "line_total_inc_vat": 1,
            "purchased_by_user_id": 1,
            "allocations": 1,
            "cost_category": 1,
            "created_at": 1,
        }
    })
    
    # Count total (separate pipeline)
    count_pipeline = pipeline.copy()
    count_pipeline.append({"$count": "total"})
    count_result = await db.invoice_lines.aggregate(count_pipeline).to_list(1)
    total = count_result[0]["total"] if count_result else 0
    
    # Sort
    sort_direction = 1 if sort_dir == "asc" else -1
    pipeline.append({"$sort": {sort_by: sort_direction}})
    
    # Paginate
    skip = (page - 1) * page_size
    pipeline.append({"$skip": skip})
    pipeline.append({"$limit": page_size})
    
    items = await db.invoice_lines.aggregate(pipeline).to_list(page_size)
    
    # Enrich with names
    for item in items:
        # Supplier name
        if item.get("supplier_id"):
            supplier = await db.counterparties.find_one(
                {"id": item["supplier_id"]}, {"_id": 0, "name": 1}
            )
            item["supplier_name"] = supplier["name"] if supplier else ""
        
        # Purchaser name
        if item.get("purchased_by_user_id"):
            purchaser = await db.users.find_one(
                {"id": item["purchased_by_user_id"]}, {"_id": 0, "first_name": 1, "last_name": 1}
            )
            item["purchased_by_name"] = f"{purchaser['first_name']} {purchaser['last_name']}" if purchaser else ""
        
        # Allocations summary
        allocations = item.get("allocations", [])
        alloc_summary = []
        for a in allocations[:3]:  # Limit to 3
            if a.get("type") == "project":
                proj = await db.projects.find_one({"id": a.get("ref_id")}, {"_id": 0, "code": 1})
                alloc_summary.append(f"P:{proj['code'] if proj else '?'}:{a.get('qty')}")
            elif a.get("type") == "warehouse":
                wh = await db.warehouses.find_one({"id": a.get("ref_id")}, {"_id": 0, "code": 1})
                alloc_summary.append(f"W:{wh['code'] if wh else '?'}:{a.get('qty')}")
        item["allocation_summary"] = ", ".join(alloc_summary)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


# ── Turnover by Counterparty ─────────────────────────────────────

@router.get("/reports/turnover-by-counterparty")
async def get_turnover_by_counterparty(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str = Query("sum_total"),
    sort_dir: str = Query("desc"),
    counterparty_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: str = Query("purchases"),  # purchases, sales, all
):
    """
    Get turnover aggregated by counterparty.
    Purchases = received invoices (supplier invoices)
    Sales = issued invoices (client invoices)
    """
    # Build match stage
    match_stage = {"org_id": user["org_id"]}
    
    if type == "purchases":
        match_stage["direction"] = "Received"
    elif type == "sales":
        match_stage["direction"] = "Issued"
    
    if counterparty_id:
        match_stage["supplier_counterparty_id"] = counterparty_id
    
    if date_from:
        match_stage["issue_date"] = match_stage.get("issue_date", {})
        match_stage["issue_date"]["$gte"] = date_from
    if date_to:
        match_stage["issue_date"] = match_stage.get("issue_date", {})
        match_stage["issue_date"]["$lte"] = date_to
    
    # Aggregation pipeline
    pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$supplier_counterparty_id",
                "count_invoices": {"$sum": 1},
                "sum_subtotal": {"$sum": "$subtotal"},
                "sum_vat": {"$sum": "$vat_amount"},
                "sum_total": {"$sum": "$total"},
                "sum_paid": {"$sum": "$paid_amount"},
                "sum_remaining": {"$sum": "$remaining_amount"},
                "first_invoice_date": {"$min": "$issue_date"},
                "last_invoice_date": {"$max": "$issue_date"},
            }
        },
    ]
    
    # Count total (before pagination)
    count_pipeline = pipeline.copy()
    count_pipeline.append({"$count": "total"})
    count_result = await db.invoices.aggregate(count_pipeline).to_list(1)
    total = count_result[0]["total"] if count_result else 0
    
    # Sort
    sort_direction = 1 if sort_dir == "asc" else -1
    pipeline.append({"$sort": {sort_by: sort_direction}})
    
    # Paginate
    skip = (page - 1) * page_size
    pipeline.append({"$skip": skip})
    pipeline.append({"$limit": page_size})
    
    results = await db.invoices.aggregate(pipeline).to_list(page_size)
    
    # Enrich with counterparty names
    items = []
    for r in results:
        counterparty_id = r["_id"]
        counterparty = None
        if counterparty_id:
            counterparty = await db.counterparties.find_one(
                {"id": counterparty_id}, {"_id": 0, "id": 1, "name": 1, "eik": 1, "type": 1}
            )
        
        items.append({
            "counterparty_id": counterparty_id,
            "counterparty_name": counterparty["name"] if counterparty else "(Неизвестен)",
            "counterparty_eik": counterparty.get("eik") if counterparty else None,
            "counterparty_type": counterparty.get("type") if counterparty else None,
            "count_invoices": r["count_invoices"],
            "sum_subtotal": round(r["sum_subtotal"] or 0, 2),
            "sum_vat": round(r["sum_vat"] or 0, 2),
            "sum_total": round(r["sum_total"] or 0, 2),
            "sum_paid": round(r["sum_paid"] or 0, 2),
            "sum_remaining": round(r["sum_remaining"] or 0, 2),
            "first_invoice_date": r["first_invoice_date"],
            "last_invoice_date": r["last_invoice_date"],
        })
    
    # Calculate grand totals
    totals_pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": None,
                "total_invoices": {"$sum": 1},
                "total_subtotal": {"$sum": "$subtotal"},
                "total_vat": {"$sum": "$vat_amount"},
                "total_amount": {"$sum": "$total"},
                "total_paid": {"$sum": "$paid_amount"},
                "total_remaining": {"$sum": "$remaining_amount"},
            }
        }
    ]
    totals_result = await db.invoices.aggregate(totals_pipeline).to_list(1)
    grand_totals = totals_result[0] if totals_result else {}
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "grand_totals": {
            "total_invoices": grand_totals.get("total_invoices", 0),
            "total_subtotal": round(grand_totals.get("total_subtotal", 0), 2),
            "total_vat": round(grand_totals.get("total_vat", 0), 2),
            "total_amount": round(grand_totals.get("total_amount", 0), 2),
            "total_paid": round(grand_totals.get("total_paid", 0), 2),
            "total_remaining": round(grand_totals.get("total_remaining", 0), 2),
        },
        "filters": {
            "type": type,
            "date_from": date_from,
            "date_to": date_to,
            "counterparty_id": counterparty_id,
        }
    }


@router.get("/reports/turnover-by-counterparty/{counterparty_id}/invoices")
async def get_counterparty_invoices(
    counterparty_id: str,
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: str = Query("purchases"),
):
    """Drill-down: Get invoices for a specific counterparty"""
    query = {
        "org_id": user["org_id"],
        "supplier_counterparty_id": counterparty_id,
    }
    
    if type == "purchases":
        query["direction"] = "Received"
    elif type == "sales":
        query["direction"] = "Issued"
    
    if date_from:
        query["issue_date"] = query.get("issue_date", {})
        query["issue_date"]["$gte"] = date_from
    if date_to:
        query["issue_date"] = query.get("issue_date", {})
        query["issue_date"]["$lte"] = date_to
    
    total = await db.invoices.count_documents(query)
    
    skip = (page - 1) * page_size
    invoices = await db.invoices.find(query, {"_id": 0}).sort("issue_date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Get counterparty name
    counterparty = await db.counterparties.find_one({"id": counterparty_id}, {"_id": 0, "name": 1})
    
    return {
        "counterparty_id": counterparty_id,
        "counterparty_name": counterparty["name"] if counterparty else "(Неизвестен)",
        "items": invoices,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }



# ── Turnover by Client (Person) ─────────────────────────────────────

@router.get("/reports/turnover-by-client")
async def get_turnover_by_client(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str = Query("sum_total"),
    sort_dir: str = Query("desc"),
    client_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """
    Get turnover aggregated by client (counterparties with type=client or person).
    Shows sales to individual clients.
    """
    # Build match stage - only Issued invoices (sales)
    match_stage = {
        "org_id": user["org_id"],
        "direction": "Issued",
    }
    
    if client_id:
        match_stage["supplier_counterparty_id"] = client_id
    
    if date_from:
        match_stage["issue_date"] = match_stage.get("issue_date", {})
        match_stage["issue_date"]["$gte"] = date_from
    if date_to:
        match_stage["issue_date"] = match_stage.get("issue_date", {})
        match_stage["issue_date"]["$lte"] = date_to
    
    # Aggregation pipeline
    pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$supplier_counterparty_id",
                "count_invoices": {"$sum": 1},
                "sum_subtotal": {"$sum": "$subtotal"},
                "sum_vat": {"$sum": "$vat_amount"},
                "sum_total": {"$sum": "$total"},
                "sum_paid": {"$sum": "$paid_amount"},
                "sum_remaining": {"$sum": "$remaining_amount"},
                "first_invoice_date": {"$min": "$issue_date"},
                "last_invoice_date": {"$max": "$issue_date"},
            }
        },
    ]
    
    # Count total
    count_pipeline = pipeline.copy()
    count_pipeline.append({"$count": "total"})
    count_result = await db.invoices.aggregate(count_pipeline).to_list(1)
    total_count = count_result[0]["total"] if count_result else 0
    
    # Sort
    sort_direction = 1 if sort_dir == "asc" else -1
    pipeline.append({"$sort": {sort_by: sort_direction}})
    
    # Paginate
    skip = (page - 1) * page_size
    pipeline.append({"$skip": skip})
    pipeline.append({"$limit": page_size})
    
    results = await db.invoices.aggregate(pipeline).to_list(page_size)
    
    # Enrich with client names (filter by type=client or person)
    items = []
    for r in results:
        client_id_val = r["_id"]
        client = None
        if client_id_val:
            client = await db.counterparties.find_one(
                {"id": client_id_val, "type": {"$in": ["client", "person", "both"]}},
                {"_id": 0, "id": 1, "name": 1, "eik": 1, "type": 1, "phone": 1, "email": 1}
            )
        
        if client:  # Only include if it's actually a client
            items.append({
                "client_id": client_id_val,
                "client_name": client["name"],
                "client_eik": client.get("eik"),
                "client_type": client.get("type"),
                "client_phone": client.get("phone"),
                "client_email": client.get("email"),
                "count_invoices": r["count_invoices"],
                "sum_subtotal": round(r["sum_subtotal"] or 0, 2),
                "sum_vat": round(r["sum_vat"] or 0, 2),
                "sum_total": round(r["sum_total"] or 0, 2),
                "sum_paid": round(r["sum_paid"] or 0, 2),
                "sum_remaining": round(r["sum_remaining"] or 0, 2),
                "first_invoice_date": r["first_invoice_date"],
                "last_invoice_date": r["last_invoice_date"],
            })
    
    # Calculate grand totals
    totals_pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": None,
                "total_invoices": {"$sum": 1},
                "total_subtotal": {"$sum": "$subtotal"},
                "total_vat": {"$sum": "$vat_amount"},
                "total_amount": {"$sum": "$total"},
                "total_paid": {"$sum": "$paid_amount"},
                "total_remaining": {"$sum": "$remaining_amount"},
            }
        }
    ]
    totals_result = await db.invoices.aggregate(totals_pipeline).to_list(1)
    grand_totals = totals_result[0] if totals_result else {}
    
    return {
        "items": items,
        "total": len(items),
        "page": page,
        "page_size": page_size,
        "total_pages": (len(items) + page_size - 1) // page_size,
        "grand_totals": {
            "total_invoices": grand_totals.get("total_invoices", 0),
            "total_subtotal": round(grand_totals.get("total_subtotal", 0), 2),
            "total_vat": round(grand_totals.get("total_vat", 0), 2),
            "total_amount": round(grand_totals.get("total_amount", 0), 2),
            "total_paid": round(grand_totals.get("total_paid", 0), 2),
            "total_remaining": round(grand_totals.get("total_remaining", 0), 2),
        },
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "client_id": client_id,
        }
    }


# ── Company Finance Summary (Dashboard) ─────────────────────────────

def get_week_number_in_month(date_str: str) -> int:
    """Get week number within a month (1-5)"""
    d = datetime.strptime(date_str, "%Y-%m-%d")
    first_day = d.replace(day=1)
    # Week number = (day + first_day_weekday - 1) // 7 + 1
    return ((d.day + first_day.weekday()) // 7) + 1


def get_week_ranges(year: int, month: int) -> list:
    """Get week ranges for a month"""
    _, days_in_month = monthrange(year, month)
    
    weeks = []
    current_week_start = 1
    
    for day in range(1, days_in_month + 1):
        current_date = datetime(year, month, day)
        # New week starts on Monday (weekday 0) or first day of month
        if current_date.weekday() == 0 and day > 1:
            weeks.append({
                "week": len(weeks) + 1,
                "start": current_week_start,
                "end": day - 1,
                "label": f"W{len(weeks) + 1}"
            })
            current_week_start = day
    
    # Add last week
    weeks.append({
        "week": len(weeks) + 1,
        "start": current_week_start,
        "end": days_in_month,
        "label": f"W{len(weeks) + 1}"
    })
    
    return weeks


@router.get("/reports/company-finance-summary")
async def get_company_finance_summary(
    user: dict = Depends(require_m5),
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    group: str = Query("week"),  # week, day, month
):
    """
    Get company finance summary with income vs expenses breakdown.
    Used for Dashboard charts.
    
    Income sources:
    - Issued invoices (sales)
    - Cash income transactions
    
    Expense sources:
    - Received invoices (purchases)
    - Cash expense transactions
    - Overhead costs
    - Payroll payments
    - Bonus payments
    """
    org_id = user["org_id"]
    
    # Date range for the month
    _, days_in_month = monthrange(year, month)
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{days_in_month:02d}"
    
    # Get week ranges
    weeks = get_week_ranges(year, month)
    
    # Initialize weekly data
    weekly_data = {w["week"]: {
        "week": w["week"],
        "label": w["label"],
        "date_range": f"{year}-{month:02d}-{w['start']:02d} - {year}-{month:02d}-{w['end']:02d}",
        "income": 0,
        "income_invoices": 0,
        "income_cash": 0,
        "expenses": 0,
        "expenses_invoices": 0,
        "expenses_cash": 0,
        "expenses_overhead": 0,
        "expenses_payroll": 0,
        "expenses_bonus": 0,
    } for w in weeks}
    
    # 1. Issued Invoices (Income)
    issued_invoices = await db.invoices.find({
        "org_id": org_id,
        "direction": "Issued",
        "issue_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "issue_date": 1, "total": 1}).to_list(1000)
    
    for inv in issued_invoices:
        week_num = get_week_number_in_month(inv["issue_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["income_invoices"] += inv.get("total", 0)
            weekly_data[week_num]["income"] += inv.get("total", 0)
    
    # 2. Received Invoices (Expenses)
    received_invoices = await db.invoices.find({
        "org_id": org_id,
        "direction": "Received",
        "issue_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "issue_date": 1, "total": 1}).to_list(1000)
    
    for inv in received_invoices:
        week_num = get_week_number_in_month(inv["issue_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_invoices"] += inv.get("total", 0)
            weekly_data[week_num]["expenses"] += inv.get("total", 0)
    
    # 3. Cash Transactions
    cash_txns = await db.cash_transactions.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "type": 1, "amount": 1}).to_list(1000)
    
    for txn in cash_txns:
        week_num = get_week_number_in_month(txn["date"])
        if week_num in weekly_data:
            if txn["type"] == "income":
                weekly_data[week_num]["income_cash"] += txn.get("amount", 0)
                weekly_data[week_num]["income"] += txn.get("amount", 0)
            else:
                weekly_data[week_num]["expenses_cash"] += txn.get("amount", 0)
                weekly_data[week_num]["expenses"] += txn.get("amount", 0)
    
    # 4. Overhead Transactions
    overhead_txns = await db.overhead_transactions.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "amount": 1}).to_list(1000)
    
    for txn in overhead_txns:
        week_num = get_week_number_in_month(txn["date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_overhead"] += txn.get("amount", 0)
            weekly_data[week_num]["expenses"] += txn.get("amount", 0)
    
    # 5. Payroll Payments (from hr/payroll system)
    payroll_payments = await db.payroll_payments.find({
        "org_id": org_id,
        "payment_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "payment_date": 1, "net_salary": 1}).to_list(1000)
    
    for pay in payroll_payments:
        week_num = get_week_number_in_month(pay["payment_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_payroll"] += pay.get("net_salary", 0)
            weekly_data[week_num]["expenses"] += pay.get("net_salary", 0)
    
    # 6. Bonus Payments
    bonus_payments = await db.bonus_payments.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "amount": 1}).to_list(1000)
    
    for bonus in bonus_payments:
        week_num = get_week_number_in_month(bonus["date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_bonus"] += bonus.get("amount", 0)
            weekly_data[week_num]["expenses"] += bonus.get("amount", 0)
    
    # Convert to list and round values
    weekly_list = []
    for w in weeks:
        data = weekly_data[w["week"]]
        for key in data:
            if isinstance(data[key], float):
                data[key] = round(data[key], 2)
        weekly_list.append(data)
    
    # Calculate totals
    total_income = sum(w["income"] for w in weekly_list)
    total_expenses = sum(w["expenses"] for w in weekly_list)
    net_balance = total_income - total_expenses
    
    # Expense breakdown for pie chart
    expense_breakdown = {
        "invoices": round(sum(w["expenses_invoices"] for w in weekly_list), 2),
        "cash": round(sum(w["expenses_cash"] for w in weekly_list), 2),
        "overhead": round(sum(w["expenses_overhead"] for w in weekly_list), 2),
        "payroll": round(sum(w["expenses_payroll"] for w in weekly_list), 2),
        "bonus": round(sum(w["expenses_bonus"] for w in weekly_list), 2),
    }
    
    # Income breakdown
    income_breakdown = {
        "invoices": round(sum(w["income_invoices"] for w in weekly_list), 2),
        "cash": round(sum(w["income_cash"] for w in weekly_list), 2),
    }
    
    # Cumulative balance for line chart
    cumulative = []
    running_balance = 0
    for w in weekly_list:
        running_balance += w["income"] - w["expenses"]
        cumulative.append({
            "week": w["week"],
            "label": w["label"],
            "balance": round(running_balance, 2)
        })
    
    return {
        "year": year,
        "month": month,
        "date_from": date_from,
        "date_to": date_to,
        "weeks": weekly_list,
        "totals": {
            "income": round(total_income, 2),
            "expenses": round(total_expenses, 2),
            "net_balance": round(net_balance, 2),
        },
        "income_breakdown": income_breakdown,
        "expense_breakdown": expense_breakdown,
        "cumulative_balance": cumulative,
    }


# ── Cash Transactions CRUD ─────────────────────────────────

@router.get("/finance/cash-transactions")
async def list_cash_transactions(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    type: Optional[str] = None,  # income, expense
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """List cash transactions"""
    query = {"org_id": user["org_id"]}
    if type:
        query["type"] = type
    if date_from:
        query["date"] = query.get("date", {})
        query["date"]["$gte"] = date_from
    if date_to:
        query["date"] = query.get("date", {})
        query["date"]["$lte"] = date_to
    
    total = await db.cash_transactions.count_documents(query)
    skip = (page - 1) * page_size
    items = await db.cash_transactions.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/finance/cash-transactions", status_code=201)
async def create_cash_transaction(
    data: dict,
    user: dict = Depends(require_m5),
):
    """Create a cash transaction (income or expense)"""
    now = datetime.now(timezone.utc).isoformat()
    txn = {
        "id": str(__import__("uuid").uuid4()),
        "org_id": user["org_id"],
        "date": data.get("date", now[:10]),
        "type": data.get("type", "expense"),  # income or expense
        "amount": data.get("amount", 0),
        "category": data.get("category", ""),
        "description": data.get("description", ""),
        "created_by": user["id"],
        "created_at": now,
    }
    await db.cash_transactions.insert_one(txn)
    return {k: v for k, v in txn.items() if k != "_id"}


# ── Overhead Transactions CRUD ─────────────────────────────────

@router.get("/finance/overhead-transactions")
async def list_overhead_transactions(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """List overhead transactions"""
    query = {"org_id": user["org_id"]}
    if date_from:
        query["date"] = query.get("date", {})
        query["date"]["$gte"] = date_from
    if date_to:
        query["date"] = query.get("date", {})
        query["date"]["$lte"] = date_to
    
    total = await db.overhead_transactions.count_documents(query)
    skip = (page - 1) * page_size
    items = await db.overhead_transactions.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/finance/overhead-transactions", status_code=201)
async def create_overhead_transaction(
    data: dict,
    user: dict = Depends(require_m5),
):
    """Create an overhead transaction"""
    now = datetime.now(timezone.utc).isoformat()
    txn = {
        "id": str(__import__("uuid").uuid4()),
        "org_id": user["org_id"],
        "date": data.get("date", now[:10]),
        "amount": data.get("amount", 0),
        "category": data.get("category", ""),
        "description": data.get("description", ""),
        "created_by": user["id"],
        "created_at": now,
    }
    await db.overhead_transactions.insert_one(txn)
    return {k: v for k, v in txn.items() if k != "_id"}


# ── Bonus Payments CRUD ─────────────────────────────────

@router.get("/finance/bonus-payments")
async def list_bonus_payments(
    user: dict = Depends(require_m5),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """List bonus payments"""
    query = {"org_id": user["org_id"]}
    if date_from:
        query["date"] = query.get("date", {})
        query["date"]["$gte"] = date_from
    if date_to:
        query["date"] = query.get("date", {})
        query["date"]["$lte"] = date_to
    
    total = await db.bonus_payments.count_documents(query)
    skip = (page - 1) * page_size
    items = await db.bonus_payments.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/finance/bonus-payments", status_code=201)
async def create_bonus_payment(
    data: dict,
    user: dict = Depends(require_m5),
):
    """Create a bonus payment"""
    now = datetime.now(timezone.utc).isoformat()
    payment = {
        "id": str(__import__("uuid").uuid4()),
        "org_id": user["org_id"],
        "date": data.get("date", now[:10]),
        "amount": data.get("amount", 0),
        "user_id": data.get("user_id"),
        "description": data.get("description", ""),
        "created_by": user["id"],
        "created_at": now,
    }
    await db.bonus_payments.insert_one(payment)
    return {k: v for k, v in payment.items() if k != "_id"}


# ── Company Finance Export (PDF/XLSX) ─────────────────────────────

from fastapi.responses import Response
from io import BytesIO

MONTH_NAMES_BG = [
    "Януари", "Февруари", "Март", "Април", "Май", "Юни",
    "Юли", "Август", "Септември", "Октомври", "Ноември", "Декември"
]


@router.get("/reports/company-finance-export")
async def export_company_finance(
    user: dict = Depends(require_m5),
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    format: str = Query("pdf", regex="^(pdf|xlsx)$"),
):
    """
    Export company finance summary as PDF or XLSX.
    Used for monthly closing reports.
    """
    # Get the finance summary data first
    org_id = user["org_id"]
    
    # Date range for the month
    _, days_in_month = monthrange(year, month)
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{days_in_month:02d}"
    
    # Get week ranges
    weeks = get_week_ranges(year, month)
    
    # Initialize weekly data
    weekly_data = {w["week"]: {
        "week": w["week"],
        "label": w["label"],
        "date_range": f"{year}-{month:02d}-{w['start']:02d} - {year}-{month:02d}-{w['end']:02d}",
        "income": 0,
        "income_invoices": 0,
        "income_cash": 0,
        "expenses": 0,
        "expenses_invoices": 0,
        "expenses_cash": 0,
        "expenses_overhead": 0,
        "expenses_payroll": 0,
        "expenses_bonus": 0,
    } for w in weeks}
    
    # Fetch data (same as company-finance-summary)
    # 1. Issued Invoices (Income)
    issued_invoices = await db.invoices.find({
        "org_id": org_id,
        "direction": "Issued",
        "issue_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "issue_date": 1, "total": 1}).to_list(1000)
    
    for inv in issued_invoices:
        week_num = get_week_number_in_month(inv["issue_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["income_invoices"] += inv.get("total", 0)
            weekly_data[week_num]["income"] += inv.get("total", 0)
    
    # 2. Received Invoices (Expenses)
    received_invoices = await db.invoices.find({
        "org_id": org_id,
        "direction": "Received",
        "issue_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "issue_date": 1, "total": 1}).to_list(1000)
    
    for inv in received_invoices:
        week_num = get_week_number_in_month(inv["issue_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_invoices"] += inv.get("total", 0)
            weekly_data[week_num]["expenses"] += inv.get("total", 0)
    
    # 3. Cash Transactions
    cash_txns = await db.cash_transactions.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "type": 1, "amount": 1}).to_list(1000)
    
    for txn in cash_txns:
        week_num = get_week_number_in_month(txn["date"])
        if week_num in weekly_data:
            if txn["type"] == "income":
                weekly_data[week_num]["income_cash"] += txn.get("amount", 0)
                weekly_data[week_num]["income"] += txn.get("amount", 0)
            else:
                weekly_data[week_num]["expenses_cash"] += txn.get("amount", 0)
                weekly_data[week_num]["expenses"] += txn.get("amount", 0)
    
    # 4. Overhead Transactions
    overhead_txns = await db.overhead_transactions.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "amount": 1}).to_list(1000)
    
    for txn in overhead_txns:
        week_num = get_week_number_in_month(txn["date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_overhead"] += txn.get("amount", 0)
            weekly_data[week_num]["expenses"] += txn.get("amount", 0)
    
    # 5. Payroll Payments
    payroll_payments = await db.payroll_payments.find({
        "org_id": org_id,
        "payment_date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "payment_date": 1, "net_salary": 1}).to_list(1000)
    
    for pay in payroll_payments:
        week_num = get_week_number_in_month(pay["payment_date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_payroll"] += pay.get("net_salary", 0)
            weekly_data[week_num]["expenses"] += pay.get("net_salary", 0)
    
    # 6. Bonus Payments
    bonus_payments_data = await db.bonus_payments.find({
        "org_id": org_id,
        "date": {"$gte": date_from, "$lte": date_to}
    }, {"_id": 0, "date": 1, "amount": 1}).to_list(1000)
    
    for bonus in bonus_payments_data:
        week_num = get_week_number_in_month(bonus["date"])
        if week_num in weekly_data:
            weekly_data[week_num]["expenses_bonus"] += bonus.get("amount", 0)
            weekly_data[week_num]["expenses"] += bonus.get("amount", 0)
    
    # Convert to list
    weekly_list = []
    for w in weeks:
        data = weekly_data[w["week"]]
        for key in data:
            if isinstance(data[key], float):
                data[key] = round(data[key], 2)
        weekly_list.append(data)
    
    # Calculate totals
    total_income = sum(w["income"] for w in weekly_list)
    total_expenses = sum(w["expenses"] for w in weekly_list)
    net_balance = total_income - total_expenses
    
    # Expense breakdown
    expense_breakdown = {
        "invoices": round(sum(w["expenses_invoices"] for w in weekly_list), 2),
        "cash": round(sum(w["expenses_cash"] for w in weekly_list), 2),
        "overhead": round(sum(w["expenses_overhead"] for w in weekly_list), 2),
        "payroll": round(sum(w["expenses_payroll"] for w in weekly_list), 2),
        "bonus": round(sum(w["expenses_bonus"] for w in weekly_list), 2),
    }
    
    # Get org info
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0, "name": 1})
    org_name = org.get("name", "Unknown") if org else "Unknown"
    
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    month_name = MONTH_NAMES_BG[month - 1]
    
    if format == "xlsx":
        return await generate_xlsx_export(
            org_name, year, month, month_name, weekly_list, 
            total_income, total_expenses, net_balance, expense_breakdown, generated_at
        )
    else:
        return await generate_pdf_export(
            org_name, year, month, month_name, weekly_list,
            total_income, total_expenses, net_balance, expense_breakdown, generated_at
        )


async def generate_xlsx_export(org_name, year, month, month_name, weekly_list, 
                               total_income, total_expenses, net_balance, expense_breakdown, generated_at):
    """Generate XLSX export"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not installed")
    
    wb = openpyxl.Workbook()
    
    # ═══ Summary Sheet ═══
    ws = wb.active
    ws.title = "Обобщение"
    
    # Header
    ws["A1"] = f"Финансов отчет - {month_name} {year}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Организация: {org_name}"
    ws["A3"] = f"Генериран на: {generated_at}"
    
    # Totals
    ws["A5"] = "ОБОБЩЕНИЕ"
    ws["A5"].font = Font(bold=True, size=12)
    
    ws["A6"] = "Общо приходи:"
    ws["B6"] = f"{total_income:.2f} лв."
    ws["B6"].font = Font(color="008000")
    
    ws["A7"] = "Общо разходи:"
    ws["B7"] = f"{total_expenses:.2f} лв."
    ws["B7"].font = Font(color="FF0000")
    
    ws["A8"] = "Нетен баланс:"
    ws["B8"] = f"{net_balance:.2f} лв."
    ws["B8"].font = Font(bold=True, color="0000FF" if net_balance >= 0 else "FF0000")
    
    # ═══ Weekly Sheet ═══
    ws2 = wb.create_sheet("По седмици")
    headers = ["Седмица", "Период", "Приходи", "Разходи", "Нето"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for row_idx, week in enumerate(weekly_list, 2):
        ws2.cell(row=row_idx, column=1, value=week["label"])
        ws2.cell(row=row_idx, column=2, value=week["date_range"])
        ws2.cell(row=row_idx, column=3, value=week["income"])
        ws2.cell(row=row_idx, column=4, value=week["expenses"])
        ws2.cell(row=row_idx, column=5, value=week["income"] - week["expenses"])
    
    # Total row
    total_row = len(weekly_list) + 2
    ws2.cell(row=total_row, column=1, value="ОБЩО").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=total_income).font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=total_expenses).font = Font(bold=True)
    ws2.cell(row=total_row, column=5, value=net_balance).font = Font(bold=True)
    
    # ═══ Breakdown Sheet ═══
    ws3 = wb.create_sheet("Разбивка разходи")
    ws3["A1"] = "Категория"
    ws3["B1"] = "Сума (лв.)"
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)
    
    breakdown_labels = {
        "invoices": "Фактури",
        "cash": "Каса",
        "overhead": "Режийни",
        "payroll": "Заплати",
        "bonus": "Бонуси"
    }
    
    for row_idx, (key, value) in enumerate(expense_breakdown.items(), 2):
        ws3.cell(row=row_idx, column=1, value=breakdown_labels.get(key, key))
        ws3.cell(row=row_idx, column=2, value=value)
    
    total_row = len(expense_breakdown) + 2
    ws3.cell(row=total_row, column=1, value="ОБЩО").font = Font(bold=True)
    ws3.cell(row=total_row, column=2, value=total_expenses).font = Font(bold=True)
    
    # Adjust column widths
    for ws_item in [ws, ws2, ws3]:
        for column_cells in ws_item.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws_item.column_dimensions[column_cells[0].column_letter].width = max_length + 2
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"finance_report_{year}_{month:02d}.xlsx"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


async def generate_pdf_export(org_name, year, month, month_name, weekly_list,
                              total_income, total_expenses, net_balance, expense_breakdown, generated_at):
    """Generate PDF export using reportlab"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab library not installed")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.gray)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceBefore=15, spaceAfter=5)
    
    story = []
    
    # Title
    story.append(Paragraph(f"Финансов отчет - {month_name} {year}", title_style))
    story.append(Paragraph(f"Организация: {org_name}", subtitle_style))
    story.append(Paragraph(f"Генериран на: {generated_at}", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Summary table
    story.append(Paragraph("ОБОБЩЕНИЕ", heading_style))
    summary_data = [
        ["Общо приходи:", f"{total_income:.2f} лв."],
        ["Общо разходи:", f"{total_expenses:.2f} лв."],
        ["Нетен баланс:", f"{net_balance:.2f} лв."],
    ]
    summary_table = Table(summary_data, colWidths=[100*mm, 60*mm])
    summary_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.green),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.red),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.blue if net_balance >= 0 else colors.red),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Weekly breakdown table
    story.append(Paragraph("ПО СЕДМИЦИ", heading_style))
    weekly_headers = ["Седмица", "Приходи", "Разходи", "Нето"]
    weekly_data = [weekly_headers]
    for week in weekly_list:
        weekly_data.append([
            week["label"],
            f"{week['income']:.2f}",
            f"{week['expenses']:.2f}",
            f"{week['income'] - week['expenses']:.2f}"
        ])
    weekly_data.append(["ОБЩО", f"{total_income:.2f}", f"{total_expenses:.2f}", f"{net_balance:.2f}"])
    
    weekly_table = Table(weekly_data, colWidths=[35*mm, 40*mm, 40*mm, 40*mm])
    weekly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(weekly_table)
    story.append(Spacer(1, 20))
    
    # Expense breakdown
    story.append(Paragraph("РАЗБИВКА НА РАЗХОДИТЕ", heading_style))
    breakdown_labels = {
        "invoices": "Фактури",
        "cash": "Каса",
        "overhead": "Режийни",
        "payroll": "Заплати",
        "bonus": "Бонуси"
    }
    breakdown_data = [["Категория", "Сума (лв.)"]]
    for key, value in expense_breakdown.items():
        breakdown_data.append([breakdown_labels.get(key, key), f"{value:.2f}"])
    breakdown_data.append(["ОБЩО", f"{total_expenses:.2f}"])
    
    breakdown_table = Table(breakdown_data, colWidths=[80*mm, 50*mm])
    breakdown_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(breakdown_table)
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"finance_report_{year}_{month:02d}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
